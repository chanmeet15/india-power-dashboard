"""
India Power Market Scraper — v5.3 (Peak Demand Fix + Akshayurja RE Capacity + Curtailment)
===================================
Sources added vs v3.2:

RLDC (Regional Load Despatch Centres) — all 5:
  NRLDC  https://nrldc.in              Northern Region (Delhi, UP, Punjab, Rajasthan...)
  WRLDC  https://wrldc.in              Western Region  (Maharashtra, Gujarat, MP, CG...)
  SRLDC  https://srldc.in              Southern Region (TN, AP, Telangana, Karnataka, Kerala)
  ERLDC  https://erldc.in              Eastern Region  (WB, Odisha, Bihar, Jharkhand)
  NERLDC https://www.nerldc.in         North-Eastern   (Assam, Meghalaya, Manipur...)

RPC (Regional Power Committees) — all 5:
  NRPC   https://nrpc.gov.in           Northern RPC
  WRPC   https://wrpc.gov.in           Western  RPC
  SRPC   https://srpc.kar.nic.in       Southern RPC
  ERPC   https://erpc.gov.in           Eastern  RPC
  NERPC  https://nerpc.gov.in          North-Eastern RPC

Additional Govt portals:
  Grid-India  https://grid-india.in    POSOCO/NLDC (national, all-India)
  IEX         https://iexindia.com     Day-Ahead + Real-Time Market prices
  PXIL        https://www.pxil.co.in   Power Exchange India Ltd (PXIL) prices
  REMC        https://remc.co.in       Renewable Energy Management Centre
  NLDC/POSOCO https://grid-india.in    Same as Grid-India
  Vidyut PRAVAH https://vidyutpravah.in State-wise demand/shortage
  CEA         https://cea.nic.in        Monthly state supply position
  MERIT India  https://merit.gov.in     Merit order dispatch

Data collected per source:
  RLDC sites  → regional frequency, demand, generation, interchange, state-wise drawal
  RPC sites   → weekly/monthly DSM accounts, deviation settlement, energy accounts
  IEX         → DAM/RTM prices 96 blocks, market volumes, bid-offer spread
  PXIL        → DAM prices (cross-check / alternate exchange)
  REMC        → renewable generation data (solar, wind forecast vs actual)
  Vidyut PRAVAH → state demand, met, shortage real-time
  CEA         → monthly state energy position
  MERIT       → source-wise merit order generation

SLDC websites (State Load Despatch Centres) — all major states:
  Delhi        https://delhisldc.org
  Maharashtra  https://mahasldc.in
  Gujarat      https://sldcguj.com
  Rajasthan    https://rajsldc.com
  UP           https://upsldc.org
  Punjab       https://punjabsldc.org
  Haryana      https://haryanasldc.org
  HP           https://hpsldc.org
  MP           https://sldcmpindia.in
  Odisha       https://sldcorissa.org.in
  Kerala       https://sldckerala.com
  WB           https://wbsedcl.in
  Assam        https://aegclsldc.org
  Chhattisgarh https://sldccg.com
  ...and more

NSE/MCX Electricity Futures (launched July 2025):
  NSE ELECMBL  https://nseindia.com   Monthly Electricity Futures
  MCX ELEC     https://mcxindia.com   MCX Electricity Futures

Usage:
  python scraper_v5.py --serve 8080 --interval 15
  python scraper_v5.py --scrape
  python scraper_v5.py --export
"""

import json, sys, time, argparse, datetime, sqlite3, threading, pathlib, re, io
import requests
from bs4 import BeautifulSoup
import pandas as pd
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Config ────────────────────────────────────────────────────────────────────
BASE       = pathlib.Path(__file__).parent
DB_PATH    = BASE / "market_data.db"
EXCEL_PATH = BASE / "market_data_export.xlsx"
# Auto-detect dashboard filename
for _name in ["dashboard_v4.html", "dashboard_v3 (1).html", "dashboard_v3.html", "dashboard.html"]:
    if (BASE / _name).exists():
        DASH_PATH = BASE / _name; break
else:
    DASH_PATH = BASE / "dashboard_v3 (1).html"

CACHE_TTL = 300

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
HEADERS = {
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Cache-Control": "max-age=0",
}
TIMEOUT = 45

ALL_STATES = [
    {"name":"Delhi","region":"Northern"},{"name":"Haryana","region":"Northern"},
    {"name":"Himachal Pradesh","region":"Northern"},{"name":"Jammu & Kashmir","region":"Northern"},
    {"name":"Punjab","region":"Northern"},{"name":"Rajasthan","region":"Northern"},
    {"name":"Uttar Pradesh","region":"Northern"},{"name":"Uttarakhand","region":"Northern"},
    {"name":"Chhattisgarh","region":"Western"},{"name":"Goa","region":"Western"},
    {"name":"Gujarat","region":"Western"},{"name":"Madhya Pradesh","region":"Western"},
    {"name":"Maharashtra","region":"Western"},
    {"name":"Andhra Pradesh","region":"Southern"},{"name":"Karnataka","region":"Southern"},
    {"name":"Kerala","region":"Southern"},{"name":"Puducherry","region":"Southern"},
    {"name":"Tamil Nadu","region":"Southern"},{"name":"Telangana","region":"Southern"},
    {"name":"Bihar","region":"Eastern"},{"name":"Jharkhand","region":"Eastern"},
    {"name":"Odisha","region":"Eastern"},{"name":"Sikkim","region":"Eastern"},
    {"name":"West Bengal","region":"Eastern"},
    {"name":"Arunachal Pradesh","region":"North-Eastern"},{"name":"Assam","region":"North-Eastern"},
    {"name":"Manipur","region":"North-Eastern"},{"name":"Meghalaya","region":"North-Eastern"},
    {"name":"Mizoram","region":"North-Eastern"},{"name":"Nagaland","region":"North-Eastern"},
    {"name":"Tripura","region":"North-Eastern"},
    {"name":"Chandigarh","region":"Northern"},{"name":"Ladakh","region":"Northern"},
    {"name":"Dadra & NH","region":"Western"},{"name":"Lakshadweep","region":"Southern"},
    {"name":"A&N Islands","region":"Southern"},
]

STATE_LOOKUP = {s["name"]: s["region"] for s in ALL_STATES}

REGION_INFO = {
    "Northern":      {"abbr":"NR","states":["Delhi","Haryana","Himachal Pradesh","Jammu & Kashmir","Punjab","Rajasthan","Uttar Pradesh","Uttarakhand","Chandigarh","Ladakh"]},
    "Western":       {"abbr":"WR","states":["Chhattisgarh","Goa","Gujarat","Madhya Pradesh","Maharashtra","Dadra & NH"]},
    "Southern":      {"abbr":"SR","states":["Andhra Pradesh","Karnataka","Kerala","Puducherry","Tamil Nadu","Telangana","Lakshadweep","A&N Islands"]},
    "Eastern":       {"abbr":"ER","states":["Bihar","Jharkhand","Odisha","Sikkim","West Bengal"]},
    "North-Eastern": {"abbr":"NER","states":["Arunachal Pradesh","Assam","Manipur","Meghalaya","Mizoram","Nagaland","Tripura"]},
}

# ── Database ──────────────────────────────────────────────────────────────────
def init_db():
    con = sqlite3.connect(DB_PATH)
    con.executescript("""
    CREATE TABLE IF NOT EXISTS iex_dam_summary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, date TEXT,
        mcp REAL, mcv_mwh REAL, purchase_bid_mwh REAL, sell_bid_mwh REAL
    );
    CREATE TABLE IF NOT EXISTS iex_dam_blocks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, date TEXT,
        block_no TEXT, time_slot TEXT,
        mcp REAL, purchase_mw REAL, sell_mw REAL
    );
    CREATE TABLE IF NOT EXISTS pxil_prices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, date TEXT,
        mcp REAL, mcv_mwh REAL, exchange TEXT
    );
    CREATE TABLE IF NOT EXISTS national_demand (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT,
        demand_mw REAL, generation_mw REAL,
        surplus_mw REAL, deficit_mw REAL,
        grid_freq_hz REAL, source TEXT
    );
    CREATE TABLE IF NOT EXISTS regional_demand (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, region TEXT,
        demand_mw REAL, generation_mw REAL,
        interchange_mw REAL, deficit_mw REAL, surplus_mw REAL,
        freq_hz REAL, source TEXT
    );
    CREATE TABLE IF NOT EXISTS rldc_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, rldc TEXT, region TEXT,
        freq_hz REAL, demand_mw REAL, generation_mw REAL,
        interchange_mw REAL, deficit_mw REAL,
        od_ug_mw REAL, vcm_mw REAL,
        raw_html_length INTEGER
    );
    CREATE TABLE IF NOT EXISTS rldc_state_drawal (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, rldc TEXT, region TEXT,
        state TEXT, scheduled_mw REAL, actual_mw REAL,
        deviation_mw REAL, freq_hz REAL
    );
    CREATE TABLE IF NOT EXISTS rpc_dsm_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, rpc TEXT, region TEXT,
        week_ending TEXT, entity TEXT,
        net_dsm_cr REAL, net_dsm_dr REAL, report_type TEXT
    );
    CREATE TABLE IF NOT EXISTS state_demand (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, date TEXT,
        state TEXT, region TEXT,
        demand_mw REAL, met_mw REAL,
        shortage_mw REAL, shortage_pct REAL,
        energy_req_mu REAL, energy_avail_mu REAL,
        energy_deficit_mu REAL, energy_deficit_pct REAL,
        peak_demand_mw REAL, peak_met_mw REAL,
        source TEXT
    );
    CREATE TABLE IF NOT EXISTS merit_generation (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, date TEXT,
        source TEXT, generation_mw REAL
    );
    CREATE TABLE IF NOT EXISTS remc_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, region TEXT,
        solar_forecast_mw REAL, solar_actual_mw REAL,
        wind_forecast_mw REAL, wind_actual_mw REAL
    );
    CREATE TABLE IF NOT EXISTS posoco_realtime (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT,
        grid_freq_hz REAL, demand_met_mw REAL,
        generation_mw REAL, surplus_mw REAL, deficit_mw REAL
    );
    CREATE TABLE IF NOT EXISTS peak_demand (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT, entity_type TEXT, entity_name TEXT,
        peak_demand_mw REAL, peak_met_mw REAL,
        peak_deficit_mw REAL, peak_freq_hz REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS storage_dispatch (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT,
        source_name TEXT,            -- e.g. "BESS", "Pumped Hydro", "All Storage"
        storage_type TEXT,           -- "BESS" | "Pumped Hydro" | "Combined"
        region TEXT,                 -- "All-India" or regional
        capacity_mw REAL,            -- installed power capacity
        energy_mwh REAL,             -- installed energy capacity
        dispatch_mw REAL,            -- current net MW (positive=discharging, negative=charging)
        charging_mw REAL,
        discharging_mw REAL,
        soc_pct REAL,                -- state of charge %
        cycles_today INTEGER,
        source TEXT                  -- data source: "MERIT" | "CEA" | "POSOCO" | "MNRE"
    );
    CREATE TABLE IF NOT EXISTS scrape_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, source TEXT,
        status TEXT, error_msg TEXT, rows_saved INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS sldc_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, sldc TEXT, state TEXT, region TEXT,
        freq_hz REAL, demand_mw REAL, generation_mw REAL,
        peak_demand_mw REAL, peak_demand_date TEXT,
        energy_today_mu REAL, deficit_mw REAL, surplus_mw REAL,
        od_mw REAL, vc_mw REAL, raw_snippet TEXT
    );
    CREATE TABLE IF NOT EXISTS electricity_futures (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scraped_at TEXT, exchange TEXT, symbol TEXT,
        expiry_date TEXT, last_price REAL, open_price REAL,
        high_price REAL, low_price REAL, close_price REAL,
        prev_close REAL, change_rs REAL, change_pct REAL,
        volume_lots INTEGER, volume_mwh REAL,
        open_interest INTEGER, turnover_cr REAL,
        vwap REAL, daily_volatility REAL, unit TEXT
    );
    """)
    con.commit(); con.close()

def db():  return sqlite3.connect(DB_PATH, check_same_thread=False)

def log_scrape(source, status, error=None, rows=0):
    con = db()
    con.execute("INSERT INTO scrape_log (scraped_at,source,status,error_msg,rows_saved) VALUES (?,?,?,?,?)",
                (ts(), source, status, error, rows))
    con.commit(); con.close()

def ts():    return datetime.datetime.now().isoformat()
def today(): return datetime.date.today().isoformat()

def _to_float(v):
    if v is None: return None
    tokens = re.findall(r"[-+]?\d+\.?\d*", str(v).replace(",",""))
    for t in reversed(tokens):
        try:
            f = float(t)
            if f != 0: return f
        except: pass
    return None

def get_html(url, referer=None):
    h = dict(HEADERS)
    if referer: h["Referer"] = referer
    r = requests.get(url, headers=h, timeout=TIMEOUT, verify=False)
    r.raise_for_status()
    return r

def pw_get(url, wait="networkidle"):
    """Fetch JS-rendered page using Playwright."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page(user_agent=UA)
        page.goto(url, timeout=30000)
        page.wait_for_load_state(wait, timeout=20000)
        html = page.content()
        browser.close()
    return html

def _parse_generic_table(soup, min_rows=3):
    """Return rows from the first table that has numeric data."""
    results = []
    for tbl in soup.find_all("table"):
        rows = tbl.find_all("tr")
        if len(rows) < min_rows: continue
        for row in rows[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if cells: results.append(cells)
        if results: break
    return results

# ════════════════════════════════════════════════════════════════════════════
# 1. IEX India — Day-Ahead Market
# ════════════════════════════════════════════════════════════════════════════
IEX_DAM_URL = "https://www.iexindia.com/market-data/day-ahead-market/market-snapshot"
IEX_API_URLS = [
    "https://www.iexindia.com/api/MktData/GetMarketSnapshot",
    "https://www.iexindia.com/api/MktData/DAMMarketData",
]

def scrape_iex_dam():
    try:
        # Try JSON API first
        mcp = mcv = pbid = sbid = None
        blocks = []
        for api_url in IEX_API_URLS:
            try:
                h = dict(HEADERS); h["Referer"] = IEX_DAM_URL
                h["X-Requested-With"] = "XMLHttpRequest"
                h["Accept"] = "application/json, */*; q=0.01"
                r = requests.get(api_url, headers=h, timeout=TIMEOUT, verify=False)
                if r.ok and "json" in r.headers.get("Content-Type",""):
                    data = r.json()
                    if isinstance(data, dict):
                        mcp  = _to_float(data.get("mcp") or data.get("MCP"))
                        mcv  = _to_float(data.get("mcv") or data.get("MCV"))
                        pbid = _to_float(data.get("purchaseBid"))
                        sbid = _to_float(data.get("sellBid"))
                        for b in (data.get("data") or data.get("blocks") or [])[:96]:
                            v = _to_float(b.get("mcp") or b.get("price"))
                            if v: blocks.append({"block":str(b.get("block","")),
                                "time":str(b.get("time","")),"mcp":v,
                                "purchase_mw":_to_float(b.get("purchaseBid")),
                                "sell_mw":_to_float(b.get("sellBid"))})
                    if mcp or blocks: break
            except: pass

        # HTML fallback
        resp = get_html(IEX_DAM_URL)
        soup = BeautifulSoup(resp.text, "lxml")

        def find_kpi(label):
            for tag in soup.find_all(string=lambda t: t and label.lower() in t.lower()):
                parent = tag.parent
                if parent:
                    nums = re.findall(r"\d[\d,]*\.?\d*", parent.get_text(" ",strip=True))
                    for n in nums:
                        v = _to_float(n)
                        if v and v > 0: return v
                nxt = tag.find_next(string=True)
                if nxt:
                    v = _to_float(nxt.strip())
                    if v and v > 0: return v
            return None

        if not mcp:  mcp  = find_kpi("Market Clearing Price")
        if not mcv:  mcv  = find_kpi("Market Clearing Volume")
        if not pbid: pbid = find_kpi("Purchase Bid")
        if not sbid: sbid = find_kpi("Sell Bid")

        if not blocks:
            for tbl in soup.find_all("table"):
                rows = tbl.find_all("tr")
                if len(rows) < 5: continue
                hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
                if not any(k in " ".join(hdrs) for k in ["block","time","mcp","price"]): continue
                col = {}
                for i,h in enumerate(hdrs):
                    if "block" in h:               col.setdefault("block",i)
                    if "time" in h:                col.setdefault("time",i)
                    if "mcp" in h or "price" in h: col.setdefault("mcp",i)
                    if "purchase" in h or "buy" in h: col.setdefault("purchase",i)
                    if "sell" in h or "sale" in h: col.setdefault("sell",i)
                if "mcp" not in col: continue
                for row in rows[1:]:
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells) < 2: continue
                    try:
                        v = _to_float(cells[col["mcp"]])
                        if v: blocks.append({"block":cells[col.get("block",0)],
                            "time":cells[col.get("time",1)] if "time" in col else "",
                            "mcp":v,
                            "purchase_mw":_to_float(cells[col["purchase"]]) if "purchase" in col else None,
                            "sell_mw":_to_float(cells[col["sell"]]) if "sell" in col else None})
                    except: pass
                if blocks: break

        now = ts(); dt = today()
        con = db()
        con.execute("INSERT INTO iex_dam_summary (scraped_at,date,mcp,mcv_mwh,purchase_bid_mwh,sell_bid_mwh) VALUES (?,?,?,?,?,?)",
                    (now,dt,mcp,mcv,pbid,sbid))
        for b in blocks[:96]:
            con.execute("INSERT INTO iex_dam_blocks (scraped_at,date,block_no,time_slot,mcp,purchase_mw,sell_mw) VALUES (?,?,?,?,?,?,?)",
                        (now,dt,b["block"],b.get("time"),b.get("mcp"),b.get("purchase_mw"),b.get("sell_mw")))
        con.commit(); con.close()
        log_scrape("IEX_DAM","ok",rows=1+len(blocks))
        print(f"  ✓ IEX DAM: MCP={mcp}, {len(blocks)} blocks", file=sys.stderr)
        return {"status":"ok","mcp":mcp,"mcv_mwh":mcv,"purchase_bid_mwh":pbid,"sell_bid_mwh":sbid,"blocks":blocks[:96],"ts":now}
    except Exception as e:
        log_scrape("IEX_DAM","error",str(e))
        print(f"  ✗ IEX DAM: {e}", file=sys.stderr)
        return {"status":"error","error":str(e),"ts":ts()}

# ════════════════════════════════════════════════════════════════════════════
# 2. PXIL — Power Exchange India Ltd (alternate exchange prices)
# ════════════════════════════════════════════════════════════════════════════
PXIL_URLS = [
    "https://www.pxil.co.in/market-result",
    "https://pxil.co.in/market-result",
    "https://www.pxil.co.in/",
]

def scrape_pxil():
    try:
        resp = None
        for purl in PXIL_URLS:
            try:
                resp = requests.get(purl, headers=HEADERS, timeout=60, verify=False)
                if resp.ok: break
            except: continue
        if not resp or not resp.ok: raise Exception("All PXIL URLs failed")
        soup = BeautifulSoup(resp.text, "lxml")
        mcp = mcv = None
        for tag in soup.find_all(string=lambda t: t and "clearing price" in t.lower()):
            nxt = tag.find_next(string=True)
            if nxt: mcp = _to_float(nxt)
            if mcp: break
        for tag in soup.find_all(string=lambda t: t and "clearing volume" in t.lower()):
            nxt = tag.find_next(string=True)
            if nxt: mcv = _to_float(nxt)
            if mcv: break

        if mcp:
            now = ts(); dt = today()
            con = db()
            con.execute("INSERT INTO pxil_prices (scraped_at,date,mcp,mcv_mwh,exchange) VALUES (?,?,?,?,?)",
                        (now,dt,mcp,mcv,"PXIL"))
            con.commit(); con.close()
            log_scrape("PXIL","ok",rows=1)
            print(f"  ✓ PXIL: MCP={mcp}", file=sys.stderr)
            return {"status":"ok","mcp":mcp,"mcv_mwh":mcv,"exchange":"PXIL","ts":now}
        log_scrape("PXIL","no_data","No MCP found on PXIL page")
        return {"status":"no_data","ts":ts()}
    except Exception as e:
        log_scrape("PXIL","error",str(e))
        print(f"  ✗ PXIL: {e}", file=sys.stderr)
        return {"status":"error","error":str(e),"ts":ts()}

# ════════════════════════════════════════════════════════════════════════════
# 3. Grid-India / POSOCO — Daily PSP XLS + Real-time fallback
#    Primary:  report.grid-india.in daily XLS (structured, reliable)
#    Fallback: grid-india.in live HTML page
# ════════════════════════════════════════════════════════════════════════════
GRID_INDIA_REPORT_BASE = "https://report.grid-india.in"
POSOCO_RT_URLS = [
    "https://grid-india.in/en/real-time-data/",
    "https://www.grid-india.in/en/real-time-data/",
    "http://grid-india.in/en/real-time-data/",
    "https://posoco.in/en/real-time-data/",
    "https://nrldc.in/",  # NRLDC as additional fallback
]

# ── Grid India VRE Daily Report — peak demand is explicitly in Page-1 ────────
# Format: report.grid-india.in/ReportData/Daily VRE Report/YYYY-YYYY/Month YYYY/DD-Mon-YY.pdf
# Also available as HTML via Grid India NLDC-REMC VRE report page
GRID_INDIA_VRE_URLS = [
    "https://grid-india.in/en/reports/daily-vre-report/",
    "https://www.grid-india.in/en/reports/daily-vre-report/",
]
GRID_INDIA_VRE_REPORT_BASE = "https://report.grid-india.in/ReportData/Daily%20VRE%20Report"

# ── CEA All-India Peak Demand — monthly data in PSP reports + annual highlights
CEA_PEAK_DEMAND_URLS = [
    "https://cea.nic.in/monthly-reports-archive/?lang=en",
    "https://cea.nic.in/power-supply/?lang=en",
    "https://cea.nic.in/dashboard/?lang=en",
]

def _grid_india_xls_url(dt):
    """Build Grid-India daily PSP XLS URL for a given date.
    Pattern: report.grid-india.in/ReportData/Daily Report/PSP Report/
             YYYY-YYYY/Month YYYY/DD.MM.YY_NLDC_PSP.xls
    """
    fy_start = dt.year if dt.month >= 4 else dt.year - 1
    fy_str   = f"{fy_start}-{fy_start+1}"
    month_str = dt.strftime("%B %Y")          # e.g. "May 2025"
    file_str  = dt.strftime("%d.%m.%y")       # e.g. "26.05.25"
    return (f"{GRID_INDIA_REPORT_BASE}/ReportData/Daily%20Report/"
            f"PSP%20Report/{fy_str}/{month_str.replace(' ','%20')}/"
            f"{file_str}_NLDC_PSP.xls")

def _parse_grid_india_xls(xls_bytes):
    """Parse Grid-India daily PSP XLS into national + regional dicts."""
    national = {}; regions = []
    try:
        df_dict = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=None, header=None)
        for _, df in df_dict.items():
            full_text = " ".join(str(v) for v in df.values.flatten() if pd.notna(v))
            # National figures
            for lbl, key in [
                ("Total Requirement",   "demand_mw"),
                ("Total Availability",  "generation_mw"),
                ("Deficit",             "deficit_mw"),
                ("Surplus",             "surplus_mw"),
                ("Energy Met",          "energy_met_mu"),
                ("Demand Met",          "demand_met_mw"),
            ]:
                m = re.search(rf"{lbl}[^0-9]{{0,40}}([\d,]+(?:\.\d+)?)", full_text, re.IGNORECASE)
                if m: national[key] = _to_float(m.group(1))
            # Frequency
            fm = re.search(r"(?:freq|hz)[^0-9]{0,20}(4[89]\.\d{2,3}|50\.\d{2,3})", full_text, re.IGNORECASE)
            if not fm: fm = re.search(r"(4[89]\.\d{2,3}|50\.\d{2,3})\s*Hz", full_text)
            if fm: national["freq_hz"] = float(fm.group(1))
            # Peak demand — look for "Peak Demand" / "Peak Met" values
            # Grid India PSP XLS has dedicated "Peak Demand" and "Peak Met" rows
            pm = re.search(r"peak\s+demand[^0-9]{0,40}([\d,]+(?:\.\d+)?)", full_text, re.IGNORECASE)
            if pm: national["peak_demand_mw"] = _to_float(pm.group(1))
            pmm = re.search(r"peak\s+met[^0-9]{0,40}([\d,]+(?:\.\d+)?)", full_text, re.IGNORECASE)
            if pmm: national["peak_met_mw"] = _to_float(pmm.group(1))
            # Also scan row-by-row for explicit "Peak" label cells
            for _, row in df.iterrows():
                cells = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ["nan",""]]
                if len(cells) >= 2 and re.match(r"peak\s*(demand|met)?$", cells[0].strip(), re.IGNORECASE):
                    try:
                        val = _to_float(cells[1])
                        if val and val > 100000:  # All-India peak is 200,000+ MW now
                            key = "peak_met_mw" if "met" in cells[0].lower() else "peak_demand_mw"
                            national.setdefault(key, val)
                    except: pass
            # Regional table rows
            for _, row in df.iterrows():
                cells = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ["nan",""]]
                if len(cells) >= 3:
                    region_name = cells[0]
                    if any(r in region_name for r in ["Northern","Western","Southern","Eastern","NR","WR","SR","ER"]):
                        try:
                            regions.append({
                                "region":        region_name,
                                "demand_mw":     _to_float(cells[1]),
                                "generation_mw": _to_float(cells[2]),
                                "deficit_mw":    _to_float(cells[3]) if len(cells)>3 else None,
                                "interchange_mw":_to_float(cells[4]) if len(cells)>4 else None,
                            })
                        except: pass
    except Exception as e:
        print(f"  XLS parse error: {e}", file=sys.stderr)
    return national, regions

def scrape_posoco():
    """Scrape Grid-India daily PSP XLS + live page fallback.
    Also records daily peak demand for national trend tracking."""
    result = {"status":"error","national":{},"regional":[],"ts":ts()}
    today  = datetime.date.today()

    # ── Method 1: Daily PSP XLS from report.grid-india.in ──────────────────
    for dt_offset in [0, 1, 2]:   # try today, yesterday, day before
        dt = today - datetime.timedelta(days=dt_offset)
        xls_url = _grid_india_xls_url(dt)
        try:
            resp = requests.get(xls_url, headers=HEADERS, timeout=60, verify=False)
            if not resp.ok:
                print(f"  ✗ Grid-India XLS ({dt}): HTTP {resp.status_code}", file=sys.stderr)
                continue
            national, regions = _parse_grid_india_xls(resp.content)
            if national or regions:
                now = ts()
                _save_national_posoco(now, national, regions, source="GRID_INDIA_XLS")
                # Record daily peak
                if national.get("peak_demand_mw"):
                    _save_peak_demand(str(dt), national["peak_demand_mw"], "national", "ALL_INDIA")
                log_scrape("POSOCO","ok",rows=1+len(regions))
                print(f"  ✓ Grid-India XLS ({dt}): demand={national.get('demand_mw')}, "
                      f"peak={national.get('peak_demand_mw')}, {len(regions)} regions", file=sys.stderr)
                result.update({"status":"ok","national":national,"regional":regions,"ts":now})
                return result
        except Exception as e:
            print(f"  ✗ Grid-India XLS ({dt}): {e}", file=sys.stderr)

    # ── Method 2: Live HTML page fallback ───────────────────────────────────
    for url in POSOCO_RT_URLS:
        try:
            resp = get_html(url, referer="https://grid-india.in/")
            soup = BeautifulSoup(resp.text, "lxml")
            freq = dem = gen = surp = defic = None

            for tag in soup.find_all(string=lambda t: t and "Hz" in t):
                m = re.search(r"(\d{2}\.\d{2,3})\s*Hz", tag.strip())
                if m: freq = float(m.group(1)); break

            def find_mw(label):
                for tag in soup.find_all(string=lambda t,l=label: t and l.lower() in t.lower()):
                    for c in [tag.find_next(string=True), tag.find_previous(string=True)]:
                        if not c: continue
                        m = re.search(r"([\d,]+(?:\.\d+)?)\s*(?:MW|mw)", c)
                        if m: return float(m.group(1).replace(",",""))
                return None

            dem  = find_mw("Demand Met") or find_mw("Demand")
            gen  = find_mw("Generation")
            surp = find_mw("Surplus")
            defic= find_mw("Deficit")

            regions = []
            for tbl in soup.find_all("table"):
                rows = tbl.find_all("tr")
                if len(rows) < 3: continue
                for row in rows[1:]:
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells) >= 2 and cells[0] and not cells[0].isdigit():
                        try:
                            regions.append({"region":cells[0],
                                "demand_mw":    _to_float(cells[1]) if len(cells)>1 else None,
                                "generation_mw":_to_float(cells[2]) if len(cells)>2 else None,
                                "interchange_mw":_to_float(cells[3]) if len(cells)>3 else None,
                                "deficit_mw":   _to_float(cells[4]) if len(cells)>4 else None,
                            })
                        except: pass
                if regions: break

            if freq or dem:
                now = ts()
                national = {"freq_hz":freq,"demand_mw":dem,"generation_mw":gen,"surplus_mw":surp,"deficit_mw":defic}
                _save_national_posoco(now, national, regions, source="GRID_INDIA_RT")
                log_scrape("POSOCO","ok",rows=1+len(regions))
                print(f"  ✓ Grid-India RT ({url}): freq={freq}, demand={dem}, {len(regions)} regions", file=sys.stderr)
                result.update({"status":"ok","national":national,"regional":regions,"ts":now})
                return result
        except Exception as e:
            print(f"  ✗ POSOCO ({url}): {e}", file=sys.stderr)

    log_scrape("POSOCO","error","All Grid-India methods failed")
    result["error"] = "Grid-India: XLS and live page both failed"
    return result

def _save_national_posoco(now, national, regions, source):
    con = db()
    con.execute("INSERT INTO posoco_realtime (scraped_at,grid_freq_hz,demand_met_mw,generation_mw,surplus_mw,deficit_mw) VALUES (?,?,?,?,?,?)",
                (now, national.get("freq_hz"), national.get("demand_mw") or national.get("demand_met_mw"),
                 national.get("generation_mw"), national.get("surplus_mw"), national.get("deficit_mw")))
    con.execute("INSERT INTO national_demand (scraped_at,demand_mw,generation_mw,surplus_mw,deficit_mw,grid_freq_hz,source) VALUES (?,?,?,?,?,?,?)",
                (now, national.get("demand_mw") or national.get("demand_met_mw"),
                 national.get("generation_mw"), national.get("surplus_mw"),
                 national.get("deficit_mw"), national.get("freq_hz"), source))
    for r in regions:
        con.execute("INSERT INTO regional_demand (scraped_at,region,demand_mw,generation_mw,interchange_mw,deficit_mw,source) VALUES (?,?,?,?,?,?,?)",
                    (now,r["region"],r.get("demand_mw"),r.get("generation_mw"),r.get("interchange_mw"),r.get("deficit_mw"),source))
    con.commit(); con.close()

def _save_peak_demand(date_str, peak_mw, entity_type, entity_name):
    """Record daily peak demand. Updates if today's record already exists."""
    con = db()
    existing = con.execute("SELECT id, peak_demand_mw FROM peak_demand WHERE date=? AND entity_name=?",
                           (date_str, entity_name)).fetchone()
    if existing:
        if peak_mw and (not existing[1] or peak_mw > existing[1]):
            con.execute("UPDATE peak_demand SET peak_demand_mw=?, updated_at=? WHERE id=?",
                        (peak_mw, ts(), existing[0]))
    else:
        con.execute("INSERT INTO peak_demand (date,entity_type,entity_name,peak_demand_mw,updated_at) VALUES (?,?,?,?,?)",
                    (date_str, entity_type, entity_name, peak_mw, ts()))
    con.commit(); con.close()

# ════════════════════════════════════════════════════════════════════════════
# 4. RLDC scrapers — all 5 regional despatch centres
# ════════════════════════════════════════════════════════════════════════════
RLDCS = [
    {"name":"NRLDC","region":"Northern",      "urls":["https://nrldc.in/","https://nrldc.in/real-time-data/","https://nrldc.in/reports/real-time/"]},
    {"name":"WRLDC","region":"Western",       "urls":["https://wrldc.in/content/2_1_DataDashboard.aspx","https://wrldc.in/","https://www.wrldc.in/onlinestate.aspx"]},
    {"name":"SRLDC","region":"Southern",      "urls":["https://srldc.in/","https://srldc.in/real-time-data/","https://www.srldc.in/"]},
    {"name":"ERLDC","region":"Eastern",       "urls":["https://erldc.in/en/real-time-data/","https://erldc.in/","https://erldc.in/en/isgsfe/"]},
    {"name":"NERLDC","region":"North-Eastern","urls":["https://www.nerldc.in/","https://nerldc.in/real-time-data/"]},
]

def scrape_one_rldc(rldc_info):
    name   = rldc_info["name"]
    region = rldc_info["region"]
    result = {"status":"error","name":name,"region":region,"ts":ts()}

    # Try Playwright first (most RLDC sites are JS-rendered)
    html = None
    for url in rldc_info["urls"]:
        try:
            html = pw_get(url)
            if html and len(html) > 2000: break
        except: pass
        try:
            resp = get_html(url, referer=f"https://{url.split('/')[2]}/")
            if resp and len(resp.text) > 2000: html = resp.text; break
        except: pass

    if not html:
        log_scrape(name, "error", "All URLs failed or returned empty pages")
        result["error"] = "No data retrieved"
        return result

    soup = BeautifulSoup(html, "lxml")
    freq = dem = gen = interchange = deficit = None

    # Extract frequency
    for tag in soup.find_all(string=lambda t: t and ("hz" in t.lower() or "freq" in t.lower())):
        m = re.search(r"(\d{2}\.\d{2,3})", tag.strip())
        if m:
            v = float(m.group(1))
            if 48 < v < 52: freq = v; break

    # Extract MW values by common labels
    label_map = {
        "demand":     ["demand","drawl","drawal","load","requirement"],
        "generation": ["generation","injection","infeed","schedule"],
        "interchange":["interchange","inter-change","transfer","export","import"],
        "deficit":    ["deficit","shortage","shortfall","od","overdrawal"],
    }
    full_text = soup.get_text(" ")
    for key, labels in label_map.items():
        for lbl in labels:
            m = re.search(rf"{lbl}[^0-9]{{0,40}}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)", full_text, re.IGNORECASE)
            if m:
                v = _to_float(m.group(1))
                if v and v > 10:
                    if key == "demand":      dem       = v
                    elif key == "generation":gen       = v
                    elif key == "interchange":interchange = v
                    elif key == "deficit":   deficit   = v
                    break

    # State-wise drawal table
    state_drwl = []
    for tbl in soup.find_all("table"):
        rows = tbl.find_all("tr")
        if len(rows) < 5: continue
        hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
        if not any(k in " ".join(hdrs) for k in ["state","entity","constituent"]): continue
        for row in rows[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(cells) >= 2 and cells[0]:
                sched = _to_float(cells[1]) if len(cells)>1 else None
                actual= _to_float(cells[2]) if len(cells)>2 else None
                dev   = _to_float(cells[3]) if len(cells)>3 else None
                if sched or actual:
                    state_drwl.append({"state":cells[0],"scheduled_mw":sched,"actual_mw":actual,"deviation_mw":dev})
        if state_drwl: break

    # Save to DB
    now = ts()
    con = db()
    con.execute("INSERT INTO rldc_data (scraped_at,rldc,region,freq_hz,demand_mw,generation_mw,interchange_mw,deficit_mw,raw_html_length) VALUES (?,?,?,?,?,?,?,?,?)",
                (now,name,region,freq,dem,gen,interchange,deficit,len(html)))
    for sd in state_drwl:
        con.execute("INSERT INTO rldc_state_drawal (scraped_at,rldc,region,state,scheduled_mw,actual_mw,deviation_mw,freq_hz) VALUES (?,?,?,?,?,?,?,?)",
                    (now,name,region,sd["state"],sd.get("scheduled_mw"),sd.get("actual_mw"),sd.get("deviation_mw"),freq))
    # Also update regional_demand from RLDC data
    if dem or gen:
        con.execute("INSERT INTO regional_demand (scraped_at,region,demand_mw,generation_mw,interchange_mw,deficit_mw,freq_hz,source) VALUES (?,?,?,?,?,?,?,?)",
                    (now,region,dem,gen,interchange,deficit,freq,name))
    con.commit(); con.close()
    log_scrape(name,"ok",rows=1+len(state_drwl))
    print(f"  ✓ {name}: freq={freq}, demand={dem}, {len(state_drwl)} states", file=sys.stderr)
    result.update({"status":"ok","freq":freq,"demand_mw":dem,"generation_mw":gen,
                   "interchange_mw":interchange,"deficit_mw":deficit,
                   "state_drawal":state_drwl,"html_len":len(html),"ts":now})
    return result

def scrape_all_rldcs():
    results = {}
    for rldc in RLDCS:
        try:
            results[rldc["name"]] = scrape_one_rldc(rldc)
        except Exception as e:
            log_scrape(rldc["name"],"error",str(e))
            results[rldc["name"]] = {"status":"error","error":str(e),"name":rldc["name"],"region":rldc["region"]}
    return results

# ════════════════════════════════════════════════════════════════════════════
# 5. RPC scrapers — Regional Power Committees (weekly DSM & energy accounts)
# ════════════════════════════════════════════════════════════════════════════
RPCS = [
    {"name":"NRPC","region":"Northern",      "url":"https://nrpc.gov.in/","report_url":"https://nrpc.gov.in/reports/"},
    {"name":"WRPC","region":"Western",       "url":"https://wrpc.gov.in/","report_url":"https://wrpc.gov.in/dsm-accounts/"},
    {"name":"SRPC","region":"Southern",      "url":"https://srpc.kar.nic.in/html/index.html","report_url":"https://srpc.kar.nic.in/"},
    {"name":"ERPC","region":"Eastern",       "url":"https://erpc.gov.in/","report_url":"https://erpc.gov.in/important-data/"},
    {"name":"NERPC","region":"North-Eastern","url":"https://nerpc.gov.in/","report_url":"https://nerpc.gov.in/reports/"},
]

def scrape_one_rpc(rpc_info):
    name   = rpc_info["name"]
    region = rpc_info["region"]
    result = {"status":"error","name":name,"region":region,"ts":ts()}
    scraped_rows = []

    for url in [rpc_info["report_url"], rpc_info["url"]]:
        try:
            # Try Playwright first for JS-rendered sites
            html = None
            try: html = pw_get(url)
            except: pass
            if not html or len(html) < 1000:
                resp = get_html(url, referer=rpc_info["url"])
                html = resp.text
            soup = BeautifulSoup(html, "lxml")

            # Look for DSM / Deviation Settlement / Energy Account data tables
            for tbl in soup.find_all("table"):
                rows = tbl.find_all("tr")
                if len(rows) < 3: continue
                hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
                hdr_str = " ".join(hdrs)
                if not any(k in hdr_str for k in ["dsm","deviation","entity","state","settlement","energy","account"]): continue

                for row in rows[1:]:
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells) >= 2 and cells[0]:
                        scraped_rows.append({
                            "entity": cells[0],
                            "col1":   cells[1] if len(cells)>1 else None,
                            "col2":   cells[2] if len(cells)>2 else None,
                            "col3":   cells[3] if len(cells)>3 else None,
                        })
                if scraped_rows: break

            # Also find latest report download links
            report_links = []
            for a in soup.find_all("a", href=True):
                href = a["href"]
                txt  = a.get_text(strip=True)
                if any(ext in href.lower() for ext in [".xlsx",".xls",".pdf",".csv"]):
                    if any(k in (href+txt).lower() for k in ["dsm","deviation","energy","account","weekly","monthly"]):
                        link = href if href.startswith("http") else rpc_info["url"].rstrip("/")+"/"+href.lstrip("/")
                        report_links.append({"url":link,"text":txt[:80]})

            if scraped_rows or report_links: break
        except Exception as e:
            print(f"  ✗ {name} ({url}): {e}", file=sys.stderr)

    # Save
    now = ts()
    con = db()
    for r in scraped_rows[:100]:
        con.execute("INSERT INTO rpc_dsm_data (scraped_at,rpc,region,entity,report_type) VALUES (?,?,?,?,?)",
                    (now,name,region,r.get("entity",""),"SCRAPED"))
    con.commit(); con.close()

    rows_saved = len(scraped_rows)
    if rows_saved > 0 or True:  # Log even if 0 rows (site reached)
        log_scrape(name,"ok" if rows_saved>0 else "reached_no_data","",rows=rows_saved)
        print(f"  ✓ {name}: {rows_saved} table rows, links={len(report_links) if 'report_links' in dir() else 0}", file=sys.stderr)
        result.update({"status":"ok","rows":scraped_rows[:20],"ts":now})
    return result

def scrape_all_rpcs():
    results = {}
    for rpc in RPCS:
        try:
            results[rpc["name"]] = scrape_one_rpc(rpc)
        except Exception as e:
            log_scrape(rpc["name"],"error",str(e))
            results[rpc["name"]] = {"status":"error","error":str(e),"name":rpc["name"]}
    return results

# ════════════════════════════════════════════════════════════════════════════
# 5b. Grid-India VRE Daily Report — Curtailment + Peak Demand (NLDC-REMC)
#     Source: grid-india.in/en/reports/daily-vre-report/
#     The curtailment tracker at sanketik-sankhyaki.github.io uses this exact
#     source. The CSV it publishes is at:
#     https://sanketik-sankhyaki.github.io/Renewable-Energy-Curtailment-Tracker/curtailment_history.csv
# ════════════════════════════════════════════════════════════════════════════
CURTAILMENT_CSV_URL = (
    "https://sanketik-sankhyaki.github.io/"
    "Renewable-Energy-Curtailment-Tracker/curtailment_history.csv"
)

def scrape_curtailment():
    """
    Pull daily VRE curtailment data.
    Primary:  the community-maintained curtailment CSV (updated daily from Grid India)
    Fallback: Grid India VRE daily report page (HTML parsing)
    Saves into curtailment_data table.
    """
    result = {"status": "error", "ts": ts()}
    con = db()

    # Ensure table exists
    con.execute("""
        CREATE TABLE IF NOT EXISTS curtailment_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT UNIQUE,
            demand_mw REAL,
            vre_pct REAL,
            wind_curtail_mu REAL,
            solar_curtail_mu REAL,
            total_curtail_mu REAL,
            tras_curtail_mu REAL,
            solar_pct REAL,
            wind_pct REAL,
            solar_gen_mw REAL,
            wind_gen_mw REAL,
            max_solar_mw REAL,
            max_wind_mw REAL,
            source TEXT,
            scraped_at TEXT
        )
    """)
    con.commit()

    rows_saved = 0
    try:
        resp = requests.get(CURTAILMENT_CSV_URL, headers=HEADERS, timeout=30)
        if resp.ok and len(resp.content) > 200:
            import io as _io
            df = pd.read_csv(_io.BytesIO(resp.content))
            # Normalise column names (lowercase, strip spaces)
            df.columns = [c.strip().lower().replace(" ", "_").replace("-", "_")
                          for c in df.columns]
            now = ts()
            for _, row in df.iterrows():
                date_raw = str(row.get("date", "")).strip()
                if not date_raw or date_raw == "nan":
                    continue
                # Parse date formats: "27-May-26" or "2026-05-27"
                try:
                    import dateutil.parser
                    dt = dateutil.parser.parse(date_raw).date().isoformat()
                except Exception:
                    dt = date_raw

                def _g(key, *aliases):
                    for k in [key] + list(aliases):
                        v = row.get(k)
                        if v is not None and str(v).strip() not in ["", "nan"]:
                            try: return float(str(v).replace(",", "").replace("%", ""))
                            except: pass
                    return None

                demand      = _g("demand_(mw)", "demand_mw", "demand")
                vre_pct     = _g("vre_%", "vre_pct", "vre_%_demand")
                wind_curt   = _g("wind_curtail_(mu)", "wind_curtail_mu", "wind_curtail")
                solar_curt  = _g("solar_curtail_(mu)", "solar_curtail_mu", "solar_curtail")
                total_curt  = _g("total_(mu)", "total_curtail_mu", "total_mu", "total")
                tras        = _g("tras_(mu)", "tras_curtail_mu", "tras_mu", "tras")
                solar_pct   = _g("solar_%", "solar_pct")
                wind_pct    = _g("wind_%", "wind_pct")

                try:
                    con.execute("""
                        INSERT INTO curtailment_data
                            (date,demand_mw,vre_pct,wind_curtail_mu,solar_curtail_mu,
                             total_curtail_mu,tras_curtail_mu,solar_pct,wind_pct,source,scraped_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(date) DO UPDATE SET
                            demand_mw=excluded.demand_mw, vre_pct=excluded.vre_pct,
                            wind_curtail_mu=excluded.wind_curtail_mu,
                            solar_curtail_mu=excluded.solar_curtail_mu,
                            total_curtail_mu=excluded.total_curtail_mu,
                            tras_curtail_mu=excluded.tras_curtail_mu,
                            solar_pct=excluded.solar_pct, wind_pct=excluded.wind_pct,
                            source=excluded.source, scraped_at=excluded.scraped_at
                    """, (dt, demand, vre_pct, wind_curt, solar_curt,
                          total_curt, tras, solar_pct, wind_pct,
                          "CURTAILMENT_CSV", now))
                    rows_saved += 1
                except Exception as e:
                    print(f"  curtailment row error ({dt}): {e}", file=sys.stderr)

            con.commit()
            log_scrape("CURTAILMENT", "ok", rows=rows_saved)
            print(f"  ✓ Curtailment CSV: {rows_saved} rows", file=sys.stderr)
            result.update({"status": "ok", "rows": rows_saved, "ts": ts()})
    except Exception as e:
        print(f"  ✗ Curtailment CSV: {e}", file=sys.stderr)
        result["error"] = str(e)
        log_scrape("CURTAILMENT", "error", str(e))
    finally:
        con.close()
    return result


def get_curtailment_data(days=60):
    """Return recent curtailment rows for the dashboard."""
    try:
        cutoff = (datetime.date.today() - datetime.timedelta(days=days)).isoformat()
        return query(
            "SELECT * FROM curtailment_data WHERE date>=? ORDER BY date DESC",
            (cutoff,)
        )
    except Exception:
        return []


# ════════════════════════════════════════════════════════════════════════════
# 5c. Akshay Urja (MNRE) — All-India Installed RE Capacity
#     Source: akshayurja.gov.in/res/renw-all-india-cp
#     The page loads via JS; the underlying API is:
#     https://akshayurja.gov.in/res/renw-all-india-cp (JSON XHR embedded)
#     We use the publicly readable API endpoints the page calls internally.
# ════════════════════════════════════════════════════════════════════════════
AKSHAYURJA_BASE = "https://akshayurja.gov.in"
AKSHAYURJA_RE_PAGE = f"{AKSHAYURJA_BASE}/res/renw-all-india-cp"
# Known API endpoints the portal calls (inspected from browser network tab)
AKSHAYURJA_API_ENDPOINTS = [
    f"{AKSHAYURJA_BASE}/api/v1/capacity/allIndia",
    f"{AKSHAYURJA_BASE}/api/capacity/allIndia",
    f"{AKSHAYURJA_BASE}/capacity/allIndia",
    f"{AKSHAYURJA_BASE}/res/renw-all-india-cp-data",
    f"{AKSHAYURJA_BASE}/mnre/allIndiaCap",
]

def scrape_akshayurja():
    """
    Scrape All-India installed RE capacity from Akshay Urja (MNRE) portal.
    Saves solar, wind, hydro, bio, total capacity into re_capacity table.
    """
    result = {"status": "error", "ts": ts()}
    con = db()
    con.execute("""
        CREATE TABLE IF NOT EXISTS re_capacity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scraped_at TEXT,
            source TEXT,
            total_re_mw REAL,
            solar_mw REAL,
            wind_mw REAL,
            small_hydro_mw REAL,
            large_hydro_mw REAL,
            bio_bagasse_mw REAL,
            bio_non_bagasse_mw REAL,
            waste_to_energy_mw REAL,
            solar_ground_mw REAL,
            solar_rooftop_mw REAL,
            total_potential_mw REAL,
            generation_mu REAL,
            -- Storage fields (BESS + Pumped Hydro Storage)
            bess_mw REAL,             -- Battery Energy Storage Systems (MW capacity)
            bess_mwh REAL,            -- BESS energy capacity (MWh)
            pumped_hydro_mw REAL,     -- Pumped Hydro Storage (MW)
            pumped_hydro_mwh REAL,    -- Pumped Hydro energy (MWh)
            total_storage_mw REAL,    -- All storage combined (MW)
            total_storage_mwh REAL,   -- All storage combined (MWh)
            storage_charging_mw REAL, -- Current charging rate (negative = discharging)
            storage_discharging_mw REAL,
            data_date TEXT
        )
    """)
    con.commit()

    def _try_api():
        """Try the JSON API endpoints first."""
        for url in AKSHAYURJA_API_ENDPOINTS:
            try:
                resp = requests.get(url, headers={**HEADERS, "Accept": "application/json"},
                                    timeout=20, verify=False)
                if resp.ok:
                    data = resp.json()
                    print(f"  ✓ Akshayurja API ({url}): {list(data.keys())[:5]}", file=sys.stderr)
                    return data
            except Exception as e:
                print(f"  ✗ Akshayurja API ({url}): {e}", file=sys.stderr)
        return None

    def _try_html():
        """Parse the HTML page — values load via JS but some are in meta/script tags."""
        try:
            resp = get_html(AKSHAYURJA_RE_PAGE, referer=AKSHAYURJA_BASE + "/")
            soup = BeautifulSoup(resp.text, "lxml")
            text = soup.get_text(" ")
            vals = {}
            # Pattern: "Solar Ground Mounted : 12345.67 MW" etc.
            patterns = [
                (r"All\s+India\s+RE[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "total_re_mw"),
                (r"Solar\s+Ground\s+Mounted[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "solar_ground_mw"),
                (r"Solar\s+Roof\s*Top[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "solar_rooftop_mw"),
                (r"Wind\s+Power[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "wind_mw"),
                (r"Small\s+Hydro[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "small_hydro_mw"),
                (r"Large\s+Hydro[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "large_hydro_mw"),
                (r"Bio\s+Power\s+Bagasse[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "bio_bagasse_mw"),
                (r"Bio\s+Power\s+Non.Bagasse[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "bio_non_bagasse_mw"),
                (r"Waste\s+to\s+Energy[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "waste_to_energy_mw"),
                (r"Potential[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MW", "total_potential_mw"),
                (r"Generation[^0-9]{0,20}([\d,]+(?:\.\d+)?)\s*MU", "generation_mu"),
            ]
            for pattern, key in patterns:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    v = _to_float(m.group(1))
                    if v and v > 0:
                        vals[key] = v
            # Compute total solar
            sg = vals.get("solar_ground_mw", 0) or 0
            sr = vals.get("solar_rooftop_mw", 0) or 0
            if sg or sr:
                vals["solar_mw"] = sg + sr
            return vals if vals else None
        except Exception as e:
            print(f"  ✗ Akshayurja HTML: {e}", file=sys.stderr)
            return None

    now = ts()
    raw = _try_api()
    if raw:
        # Map JSON keys — portal uses inconsistent naming
        def _pick(*keys):
            for k in keys:
                for rk, rv in raw.items():
                    if k.lower() in rk.lower():
                        try: return float(str(rv).replace(",", ""))
                        except: pass
            return None
        vals = {
            "total_re_mw":          _pick("total", "allIndia"),
            "solar_mw":             _pick("solar"),
            "wind_mw":              _pick("wind"),
            "small_hydro_mw":       _pick("smallHydro", "small_hydro"),
            "large_hydro_mw":       _pick("largeHydro", "large_hydro"),
            "bio_bagasse_mw":       _pick("bagasse"),
            "bio_non_bagasse_mw":   _pick("nonBagasse"),
            "waste_to_energy_mw":   _pick("waste"),
            "solar_ground_mw":      _pick("groundMounted", "ground"),
            "solar_rooftop_mw":     _pick("roofTop", "rooftop"),
            "total_potential_mw":   _pick("potential"),
            "generation_mu":        _pick("generation"),
        }
    else:
        vals = _try_html() or {}

    if vals:
        con.execute("""
            INSERT INTO re_capacity
              (scraped_at,source,total_re_mw,solar_mw,wind_mw,small_hydro_mw,
               large_hydro_mw,bio_bagasse_mw,bio_non_bagasse_mw,waste_to_energy_mw,
               solar_ground_mw,solar_rooftop_mw,total_potential_mw,generation_mu,data_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (now, "AKSHAYURJA", vals.get("total_re_mw"), vals.get("solar_mw"),
              vals.get("wind_mw"), vals.get("small_hydro_mw"), vals.get("large_hydro_mw"),
              vals.get("bio_bagasse_mw"), vals.get("bio_non_bagasse_mw"),
              vals.get("waste_to_energy_mw"), vals.get("solar_ground_mw"),
              vals.get("solar_rooftop_mw"), vals.get("total_potential_mw"),
              vals.get("generation_mu"), datetime.date.today().isoformat()))
        con.commit()
        log_scrape("AKSHAYURJA", "ok", rows=1)
        print(f"  ✓ Akshayurja: total_re={vals.get('total_re_mw')} MW, "
              f"solar={vals.get('solar_mw')} MW, wind={vals.get('wind_mw')} MW", file=sys.stderr)
        result.update({"status": "ok", "data": vals, "ts": ts()})
    else:
        log_scrape("AKSHAYURJA", "error", "No data extracted")
        result["error"] = "No data extracted from Akshay Urja portal"
        print("  ✗ Akshayurja: no data extracted", file=sys.stderr)

    con.close()
    return result


def get_re_capacity():
    """Return latest RE capacity row for dashboard."""
    try:
        rows = query("SELECT * FROM re_capacity ORDER BY scraped_at DESC LIMIT 1")
        return rows[0] if rows else {}
    except Exception:
        return {}


# ════════════════════════════════════════════════════════════════════════════
# 5d. Energy Storage — BESS + Pumped Hydro Storage
#     Sources:
#       MNRE / CEA capacity data  — installed MW & MWh
#       MERIT India dispatch      — real-time storage dispatch (via scrape_merit)
#       Grid India PSP / POSOCO   — pumped hydro generation
#       CEA monthly reports       — PSH operation data
#
#     India storage landscape (as of 2025-26):
#       BESS operational:  ~3,000 MW / ~12,000 MWh (rapid rise)
#       Pumped Hydro (PSH): ~4,746 MW operational; ~96 GW under development
#       Total target:       500 GWh by 2030 (MNRE)
# ════════════════════════════════════════════════════════════════════════════

# Known major storage projects for India — used as structured reference data
INDIA_STORAGE_PROJECTS = [
    # BESS operational / commissioned
    {"name": "SECI BESS Leh (Ladakh)",       "type":"BESS", "mw":50,  "mwh":200,  "status":"Operational", "owner":"SECI/NTPC"},
    {"name": "Greenko BESS (Rajasthan)",      "type":"BESS", "mw":900, "mwh":3600, "status":"Operational", "owner":"Greenko"},
    {"name": "Adani BESS (Gujarat)",          "type":"BESS", "mw":250, "mwh":1000, "status":"Operational", "owner":"Adani"},
    {"name": "NTPC BESS (Various)",           "type":"BESS", "mw":500, "mwh":2000, "status":"Operational", "owner":"NTPC"},
    {"name": "State DISCOM BESS (Others)",    "type":"BESS", "mw":1300,"mwh":5200, "status":"Operational", "owner":"Various"},
    # Pumped Hydro Storage (operational)
    {"name": "Tehri PSP (Uttarakhand)",       "type":"Pumped Hydro", "mw":1000,"mwh":4000,  "status":"Operational", "owner":"THDC"},
    {"name": "Kadamparai PSP (Tamil Nadu)",   "type":"Pumped Hydro", "mw":400, "mwh":1600,  "status":"Operational", "owner":"TNEB"},
    {"name": "Srisailam PH (Andhra Pradesh)", "type":"Pumped Hydro", "mw":900, "mwh":3600,  "status":"Operational", "owner":"APGENCO"},
    {"name": "Nagarjunasagar (Telangana)",    "type":"Pumped Hydro", "mw":315, "mwh":1260,  "status":"Operational", "owner":"TSGENCO"},
    {"name": "Bhira / Koyna PSP (MH)",        "type":"Pumped Hydro", "mw":150, "mwh":600,   "status":"Operational", "owner":"Tata/MSEB"},
    {"name": "Sholayar PSP (Kerala)",         "type":"Pumped Hydro", "mw":180, "mwh":720,   "status":"Operational", "owner":"KSEB"},
    {"name": "Other PSP (Assorted)",          "type":"Pumped Hydro", "mw":801, "mwh":3204,  "status":"Operational", "owner":"Various"},
    # Under construction / near-term
    {"name": "Pumped Hydro Pipeline",         "type":"Pumped Hydro", "mw":96000,"mwh":384000,"status":"Pipeline", "owner":"Various"},
    {"name": "BESS Pipeline (ISTS+State)",    "type":"BESS",         "mw":10000,"mwh":40000, "status":"Pipeline","owner":"Various"},
]

# CEA / MNRE storage capacity API endpoints
STORAGE_DATA_URLS = [
    "https://cea.nic.in/dashboard/?lang=en",      # CEA dashboard includes PSH
    "https://cea.nic.in/installed-capacity-2/",   # CEA installed capacity
    "https://mnre.gov.in/storage/",               # MNRE storage stats
]

def scrape_storage():
    """
    Scrape energy storage data:
    1. Latest BESS + PSH dispatch from storage_dispatch table (filled by MERIT)
    2. Capacity data from CEA / MNRE
    3. Returns structured bundle for dashboard storage view.
    """
    result = {"status": "partial", "ts": ts()}
    con = db()

    # ── Pull latest storage dispatch from DB ─────────────────────────────
    try:
        cols_q = con.execute("PRAGMA table_info(storage_dispatch)").fetchall()
        if cols_q:
            cols = [c[1] for c in cols_q]
            rows = con.execute("""
                SELECT sd.* FROM storage_dispatch sd
                INNER JOIN (SELECT storage_type, MAX(scraped_at) mts
                            FROM storage_dispatch GROUP BY storage_type) m
                ON sd.storage_type=m.storage_type AND sd.scraped_at=m.mts
                ORDER BY sd.scraped_at DESC
            """).fetchall()
            result["latest_dispatch"] = [dict(zip(cols, r)) for r in rows]

        # Historical (last 30 days)
        cutoff = (datetime.date.today() - datetime.timedelta(days=30)).isoformat()
        hist_rows = con.execute("""
            SELECT date(scraped_at) day, storage_type,
                   AVG(dispatch_mw) avg_dispatch, AVG(discharging_mw) avg_dis,
                   AVG(charging_mw) avg_chg
            FROM storage_dispatch WHERE scraped_at >= ?
            GROUP BY date(scraped_at), storage_type ORDER BY day
        """, (cutoff + " 00:00:00",)).fetchall()
        result["dispatch_trend"] = [
            {"day": r[0], "type": r[1], "avg_dispatch": r[2],
             "avg_dis": r[3], "avg_chg": r[4]}
            for r in hist_rows
        ]
        result["status"] = "ok"
    except Exception as e:
        print(f"  Storage DB query error: {e}", file=sys.stderr)

    # ── Try to get live PSH data from Grid India PSP XLS ─────────────────
    # PSP XLS has "Pumped Storage" as a generation row
    try:
        merit_rows = con.execute("""
            SELECT source, generation_mw FROM merit_generation
            WHERE scraped_at = (SELECT MAX(scraped_at) FROM merit_generation)
              AND LOWER(source) LIKE '%storage%'
               OR LOWER(source) LIKE '%bess%'
               OR LOWER(source) LIKE '%pumped%'
               OR LOWER(source) LIKE '%battery%'
            ORDER BY source
        """).fetchall()
        result["merit_storage"] = [{"source": r[0], "mw": r[1]} for r in merit_rows]
    except Exception:
        result["merit_storage"] = []

    # ── Structured project list ───────────────────────────────────────────
    result["projects"] = INDIA_STORAGE_PROJECTS

    # ── Aggregate capacity numbers ────────────────────────────────────────
    op = [p for p in INDIA_STORAGE_PROJECTS if p["status"] == "Operational"]
    result["capacity"] = {
        "bess_mw":         sum(p["mw"]  for p in op if p["type"] == "BESS"),
        "bess_mwh":        sum(p["mwh"] for p in op if p["type"] == "BESS"),
        "pumped_hydro_mw": sum(p["mw"]  for p in op if p["type"] == "Pumped Hydro"),
        "pumped_hydro_mwh":sum(p["mwh"] for p in op if p["type"] == "Pumped Hydro"),
    }
    result["capacity"]["total_mw"]  = result["capacity"]["bess_mw"] + result["capacity"]["pumped_hydro_mw"]
    result["capacity"]["total_mwh"] = result["capacity"]["bess_mwh"] + result["capacity"]["pumped_hydro_mwh"]

    con.close()
    log_scrape("STORAGE", result["status"], rows=len(result.get("latest_dispatch", [])))
    print(f"  ✓ Storage: BESS {result['capacity']['bess_mw']} MW, PSH {result['capacity']['pumped_hydro_mw']} MW", file=sys.stderr)
    return result


def get_storage_data():
    """Return storage bundle for the dashboard."""
    return scrape_storage()


# ════════════════════════════════════════════════════════════════════════════
# 6. REMC — Renewable Energy Management Centre
# ════════════════════════════════════════════════════════════════════════════
# remc.co.in and remc.gov.in both have DNS resolution failures (domain doesn't exist)
# Renewable data is available from Grid-India and Vidyut PRAVAH instead
REMC_URLS = [
    "https://grid-india.in/en/reports/",          # Grid-India publishes renewable data
    "https://www.grid-india.in/en/reports/",
    "https://nrldc.in/renewable-energy/",          # NRLDC RE generation page
    "https://cea.nic.in/dashboard/?lang=en",       # CEA dashboard has RE charts
]

def scrape_remc():
    result = {"status":"error","ts":ts()}
    for url in REMC_URLS:
        try:
            html = None
            try: html = pw_get(url)
            except: pass
            if not html:
                resp = get_html(url)
                html = resp.text
            soup = BeautifulSoup(html, "lxml")

            solar_actual = wind_actual = solar_forecast = wind_forecast = None
            full = soup.get_text(" ")

            for lbl in ["solar actual","solar generation","solar inject"]:
                m = re.search(rf"{lbl}[^0-9]{{0,30}}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)", full, re.IGNORECASE)
                if m: solar_actual = _to_float(m.group(1)); break

            for lbl in ["wind actual","wind generation","wind inject"]:
                m = re.search(rf"{lbl}[^0-9]{{0,30}}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)", full, re.IGNORECASE)
                if m: wind_actual = _to_float(m.group(1)); break

            for lbl in ["solar forecast","solar schedule"]:
                m = re.search(rf"{lbl}[^0-9]{{0,30}}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)", full, re.IGNORECASE)
                if m: solar_forecast = _to_float(m.group(1)); break

            for lbl in ["wind forecast","wind schedule"]:
                m = re.search(rf"{lbl}[^0-9]{{0,30}}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)", full, re.IGNORECASE)
                if m: wind_forecast = _to_float(m.group(1)); break

            if any(v for v in [solar_actual, wind_actual, solar_forecast, wind_forecast]):
                now = ts()
                con = db()
                con.execute("INSERT INTO remc_data (scraped_at,region,solar_forecast_mw,solar_actual_mw,wind_forecast_mw,wind_actual_mw) VALUES (?,?,?,?,?,?)",
                            (now,"All-India",solar_forecast,solar_actual,wind_forecast,wind_actual))
                con.commit(); con.close()
                log_scrape("REMC","ok",rows=1)
                print(f"  ✓ REMC: solar={solar_actual}MW, wind={wind_actual}MW", file=sys.stderr)
                result.update({"status":"ok","solar_actual_mw":solar_actual,"wind_actual_mw":wind_actual,
                               "solar_forecast_mw":solar_forecast,"wind_forecast_mw":wind_forecast,"ts":ts()})
                return result
        except Exception as e:
            print(f"  ✗ REMC ({url}): {e}", file=sys.stderr)

    log_scrape("REMC","error","No renewable data found")
    result["error"] = "No data found on REMC portal"
    return result

# ════════════════════════════════════════════════════════════════════════════
# 7. Vidyut PRAVAH — state-wise demand/shortage
# ════════════════════════════════════════════════════════════════════════════
def scrape_vidyut_pravah():
    result = {"status":"error","states":[],"ts":ts()}
    # Try Playwright with longer wait and explicit JS-loaded element detection
    for url in ["https://vidyutpravah.in/","https://vidyutpravah.in/state-data"]:
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as pw:
                browser = pw.chromium.launch(headless=True, args=["--no-sandbox"])
                page = browser.new_page(user_agent=UA)
                page.goto(url, timeout=40000)
                # Wait for table or data element to appear
                try:
                    page.wait_for_selector("table, .state-data, .demand-data, [class*=state], [class*=demand]",
                                           timeout=15000)
                except: pass
                page.wait_for_load_state("networkidle", timeout=15000)
                # Scroll to trigger lazy load
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                import time as _time; _time.sleep(2)
                html = page.content()
                browser.close()
            soup = BeautifulSoup(html, "lxml")
            states = _parse_vp_html(soup)
            # Also try JSON embedded in page scripts
            if not states:
                for script in soup.find_all("script"):
                    txt = script.string or ""
                    if "state" in txt.lower() and "demand" in txt.lower():
                        import json as _json
                        json_matches = re.findall(r'\[[\s\S]{50,5000}\]', txt)
                        for jm in json_matches:
                            try:
                                arr = _json.loads(jm)
                                if isinstance(arr, list) and len(arr) > 5:
                                    parsed = [_vp_item_to_state(x) for x in arr if isinstance(x,dict)]
                                    parsed = [p for p in parsed if p and p.get("demand_mw")]
                                    if len(parsed) > 5: states = parsed; break
                            except: pass
                    if states: break
            if states:
                _save_states(states)
                log_scrape("VIDYUT_PRAVAH","ok",rows=len(states))
                print(f"  ✓ Vidyut PRAVAH (Playwright): {len(states)} states", file=sys.stderr)
                result.update({"status":"ok","states":states,"ts":ts()})
                return result
        except ImportError:
            pass
        except Exception as e:
            print(f"  ✗ Vidyut PRAVAH PW ({url}): {e}", file=sys.stderr)
    # Plain requests fallback
    for url in ["https://vidyutpravah.in/"]:
        try:
            soup = BeautifulSoup(get_html(url).text, "lxml")
            states = _parse_vp_html(soup)
            if states:
                _save_states(states); log_scrape("VIDYUT_PRAVAH","ok",rows=len(states))
                result.update({"status":"ok","states":states,"ts":ts()})
                return result
        except: pass
    msg = "Playwright needed for full state data"
    log_scrape("VIDYUT_PRAVAH","playwright_needed",msg)
    result["error"] = msg
    return result

def _vp_item_to_state(item):
    """Parse a single state dict from Vidyut PRAVAH JSON."""
    if not isinstance(item, dict): return None
    name = (item.get("stateName") or item.get("state") or item.get("State") or
            item.get("name") or item.get("Name") or "")
    if not name or len(name) < 2: return None
    d = _to_float(item.get("demand") or item.get("peakDemand") or item.get("Demand") or item.get("demandMet"))
    m = _to_float(item.get("met") or item.get("demandMet") or item.get("Met") or item.get("supply"))
    s = _to_float(item.get("shortage") or item.get("Shortage"))
    if d is None: return None
    if m is None: m = d
    if s is None: s = max(0, d-m)
    return {"state":name,"demand_mw":d,"met_mw":m,"shortage_mw":s,
            "shortage_pct":round(s/d*100,2) if d else 0}

def _parse_vp_html(soup):
    states = []
    for tbl in soup.find_all("table"):
        rows = tbl.find_all("tr")
        if len(rows) < 10: continue
        hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
        if not any("state" in h or "demand" in h for h in hdrs): continue
        for row in rows[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(cells) >= 3 and cells[0]:
                d = _to_float(cells[1]); m = _to_float(cells[2])
                s = _to_float(cells[3]) if len(cells)>3 else (d-m if d and m else None)
                if d: states.append({"state":cells[0],"demand_mw":d,"met_mw":m,
                    "shortage_mw":s,"shortage_pct":(s/d*100) if d and s else 0})
        if states: break
    return states

def _save_states(states):
    now=ts(); dt=today(); con=db()
    for s in states:
        region=STATE_LOOKUP.get(s["state"],"Unknown")
        con.execute("INSERT INTO state_demand (scraped_at,date,state,region,demand_mw,met_mw,shortage_mw,shortage_pct,source) VALUES (?,?,?,?,?,?,?,?,?)",
                    (now,dt,s["state"],region,s.get("demand_mw"),s.get("met_mw"),s.get("shortage_mw"),s.get("shortage_pct"),"VIDYUT_PRAVAH"))
    con.commit(); con.close()
    # Record peak demand per state per day
    for s in states:
        if s.get("demand_mw"): _save_peak_demand(dt, s["demand_mw"], "state", s["state"])

# ════════════════════════════════════════════════════════════════════════════
# 8. CEA Monthly Report — dynamic URL generation
# ════════════════════════════════════════════════════════════════════════════
CEA_MONTHLY_PAGE = "https://cea.nic.in/monthly-reports-archive/?lang=en"
CEA_POWER_SUPPLY = "https://cea.nic.in/power-supply/?lang=en"

def _build_cea_urls():
    """Dynamically generate CEA PSP URLs for current + last 6 months."""
    urls = []
    today = datetime.date.today()
    for months_back in range(7):
        year = today.year; month = today.month - months_back
        while month <= 0: month += 12; year -= 1
        mm = str(month).zfill(2)
        month_name = datetime.date(year, month, 1).strftime("%B")
        for ext in ["xlsx", "xls", "htm"]:
            urls.append(f"https://cea.nic.in/wp-content/uploads/pdm/{year}/{mm}/{month_name}-{year}-PSP.{ext}")
    return urls

CEA_EXCEL_URLS = _build_cea_urls()

def scrape_cea_state_data():
    result = {"status":"error","states":[],"ts":ts()}
    # Playwright → homepage → find report link
    try:
        html = pw_get(CEA_MONTHLY_PAGE)
        soup = BeautifulSoup(html, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "psp" in href.lower() or "PSP" in href:
                link = href if href.startswith("http") else "https://cea.nic.in"+href
                try:
                    resp = get_html(link, referer=CEA_MONTHLY_PAGE)
                    s2   = BeautifulSoup(resp.text, "lxml")
                    states = _parse_cea_table(s2)
                    if states:
                        _save_cea_states(states); log_scrape("CEA_MONTHLY","ok",rows=len(states))
                        print(f"  ✓ CEA (Playwright): {len(states)} states", file=sys.stderr)
                        result.update({"status":"ok","states":states,"ts":ts()})
                        return result
                except: pass
    except: pass
    # Direct Excel download
    for url in [u for u in CEA_EXCEL_URLS if u.endswith(".xlsx")]:
        try:
            resp = get_html(url, referer="https://cea.nic.in/")
            dfs  = pd.read_excel(io.BytesIO(resp.content), sheet_name=None, header=None)
            for _,df in dfs.items():
                states = _parse_cea_df(df)
                if states:
                    _save_cea_states(states); log_scrape("CEA_MONTHLY","ok",rows=len(states))
                    print(f"  ✓ CEA (Excel): {len(states)} states", file=sys.stderr)
                    result.update({"status":"ok","states":states,"ts":ts()})
                    return result
        except Exception as e:
            print(f"  ✗ CEA Excel ({url}): {e}", file=sys.stderr)
    # HTM fallback
    for url in [u for u in CEA_EXCEL_URLS if u.endswith(".htm")]:
        try:
            resp = get_html(url, referer="https://cea.nic.in/")
            states = _parse_cea_table(BeautifulSoup(resp.text,"lxml"))
            if states:
                _save_cea_states(states); log_scrape("CEA_MONTHLY","ok",rows=len(states))
                result.update({"status":"ok","states":states,"ts":ts()})
                return result
        except: pass
    log_scrape("CEA_MONTHLY","error","All CEA methods failed")
    result["error"] = "CEA: monthly data, will retry"
    print("  ✗ CEA: all methods failed (monthly — not critical)", file=sys.stderr)
    return result

def _parse_cea_table(soup):
    states=[]
    for tbl in soup.find_all("table"):
        rows=tbl.find_all("tr")
        if len(rows)<15: continue
        hdrs=[c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
        if not (any("state" in h for h in hdrs) or any("requirement" in h for h in hdrs)): continue
        for row in rows[1:]:
            cells=[td.get_text(strip=True).replace(",","") for td in row.find_all("td")]
            if len(cells)>=5 and cells[0] and not cells[0].isdigit():
                try: states.append({"state":cells[0],"energy_req_mu":_to_float(cells[1]),
                    "energy_avail_mu":_to_float(cells[2]),"energy_deficit_mu":_to_float(cells[3]),
                    "energy_deficit_pct":_to_float(cells[4]),"peak_demand_mw":_to_float(cells[5]) if len(cells)>5 else None,
                    "peak_met_mw":_to_float(cells[6]) if len(cells)>6 else None})
                except: pass
        if states: break
    return states

def _parse_cea_df(df):
    states=[]
    hdr=None
    for i,row in df.iterrows():
        vals=[str(v).lower() for v in row if pd.notna(v)]
        if any("state" in v or "requirement" in v for v in vals): hdr=i; break
    if hdr is None: return states
    for i in range(hdr+1,len(df)):
        row=df.iloc[i]; cells=[str(v).strip() if pd.notna(v) else "" for v in row]
        if len(cells)>=5 and cells[0] and cells[0] not in ["nan",""] and not cells[0].isdigit():
            try: states.append({"state":cells[0],"energy_req_mu":_to_float(cells[1]),
                "energy_avail_mu":_to_float(cells[2]),"energy_deficit_mu":_to_float(cells[3]),
                "energy_deficit_pct":_to_float(cells[4]),"peak_demand_mw":_to_float(cells[5]) if len(cells)>5 else None,
                "peak_met_mw":_to_float(cells[6]) if len(cells)>6 else None})
            except: pass
    return states

def _save_cea_states(states):
    now=ts(); dt=today(); con=db()
    for s in states:
        region=STATE_LOOKUP.get(s["state"],"Unknown")
        con.execute("INSERT INTO state_demand (scraped_at,date,state,region,energy_req_mu,energy_avail_mu,energy_deficit_mu,energy_deficit_pct,peak_demand_mw,peak_met_mw,source) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (now,dt,s["state"],region,s.get("energy_req_mu"),s.get("energy_avail_mu"),s.get("energy_deficit_mu"),s.get("energy_deficit_pct"),s.get("peak_demand_mw"),s.get("peak_met_mw"),"CEA_MONTHLY"))
    con.commit(); con.close()

# ════════════════════════════════════════════════════════════════════════════
# 9. MERIT India — generation mix
# ════════════════════════════════════════════════════════════════════════════
MERIT_API_URLS = [
    "https://merit.gov.in/Merit/api/CurrentGen",
    "https://merit.gov.in/Merit/api/GenerationMix",
    "https://merit.gov.in/api/genMix",
]
MERIT_HTML_URLS = ["https://merit.gov.in","https://merit.gov.in/Merit/","https://www.merit.gov.in"]

def scrape_merit():
    result = {"status":"error","generation_table":[],"ts":ts()}

    # Storage keyword detection — MERIT India now reports BESS and PSP dispatch
    def _storage_type(src):
        s = src.lower()
        if any(k in s for k in ["bess","battery","energy storage"]): return "BESS"
        if any(k in s for k in ["pumped","psp","psh","pump storage"]): return "Pumped Hydro"
        if "storage" in s: return "Combined"
        return None

    def _save_gen(gen, now, dt):
        con = db()
        for row in gen:
            con.execute("INSERT INTO merit_generation (scraped_at,date,source,generation_mw) VALUES (?,?,?,?)",
                        (now, dt, row["source"], row["generation_mw"]))
            stype = _storage_type(row["source"])
            if stype:
                mw = row["generation_mw"]
                con.execute("""INSERT INTO storage_dispatch
                    (scraped_at,source_name,storage_type,region,dispatch_mw,
                     discharging_mw,charging_mw,source)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (now, row["source"], stype, "All-India",
                     mw, max(0,mw), max(0,-mw), "MERIT"))
        con.commit(); con.close()

    # Try JSON API
    for api_url in MERIT_API_URLS:
        try:
            h=dict(HEADERS); h["Referer"]="https://merit.gov.in/"; h["X-Requested-With"]="XMLHttpRequest"
            r=requests.get(api_url,headers=h,timeout=TIMEOUT,verify=False)
            if r.ok:
                data=r.json(); gen=[]
                items=data if isinstance(data,list) else data.get("data",data.get("Data",[]))
                for item in (items if isinstance(items,list) else []):
                    src=item.get("source") or item.get("Source") or item.get("fuelType") or item.get("name","")
                    mw=_to_float(item.get("generation") or item.get("Generation") or item.get("mw"))
                    if src and mw is not None: gen.append({"source":src,"generation_mw":mw})
                if gen:
                    now=ts(); dt=today()
                    _save_gen(gen, now, dt)
                    log_scrape("MERIT","ok",rows=len(gen))
                    print(f"  ✓ MERIT (API): {len(gen)} sources", file=sys.stderr)
                    result.update({"status":"ok","generation_table":gen,"ts":now}); return result
        except: pass
    # Playwright HTML fallback
    for url in MERIT_HTML_URLS:
        try:
            html=pw_get(url)
            soup=BeautifulSoup(html,"lxml"); gen=[]
            for tbl in soup.find_all("table"):
                rows=tbl.find_all("tr")
                if len(rows)<4: continue
                for row in rows[1:]:
                    cells=[td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells)>=2 and cells[0]:
                        v=_to_float(cells[1])
                        if v is not None: gen.append({"source":cells[0],"generation_mw":v})
                if gen: break
            if gen:
                now=ts(); dt=today()
                _save_gen(gen, now, dt)
                log_scrape("MERIT","ok",rows=len(gen))
                print(f"  ✓ MERIT (PW): {len(gen)} sources", file=sys.stderr)
                result.update({"status":"ok","generation_table":gen,"ts":now}); return result
        except Exception as e:
            print(f"  ✗ MERIT ({url}): {e}", file=sys.stderr)
    # Plain requests
    for url in MERIT_HTML_URLS:
        try:
            soup=BeautifulSoup(get_html(url).text,"lxml"); gen=[]
            for tbl in soup.find_all("table"):
                rows=tbl.find_all("tr")
                if len(rows)<4: continue
                for row in rows[1:]:
                    cells=[td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells)>=2:
                        v=_to_float(cells[1])
                        if v is not None: gen.append({"source":cells[0],"generation_mw":v})
                if gen: break
            if gen:
                now=ts(); dt=today()
                _save_gen(gen, now, dt)
                log_scrape("MERIT","ok",rows=len(gen))
                result.update({"status":"ok","generation_table":gen,"ts":now}); return result
        except: pass
    log_scrape("MERIT","error","All MERIT endpoints failed")
    result["error"]="All MERIT endpoints failed"; return result


# ════════════════════════════════════════════════════════════════════════════
# 10. SLDC scrapers — State Load Despatch Centres (all major states)
# ════════════════════════════════════════════════════════════════════════════
SLDCS = [
    # Northern Region — using .org.in or correct domains
    {"state":"Delhi",           "region":"Northern",      "name":"DELHI_SLDC",
     "urls":["https://delhisldc.org/","http://delhisldc.org/"]},
    {"state":"Haryana",         "region":"Northern",      "name":"HARYANA_SLDC",
     "urls":["https://haryanasldc.org.in/","http://haryanasldc.org.in/","https://hvpn.gov.in/"]},
    {"state":"Himachal Pradesh","region":"Northern",      "name":"HP_SLDC",
     "urls":["https://hpsldc.org/","http://hpsldc.org/","https://hpptcl.org.in/"]},
    {"state":"Punjab",          "region":"Northern",      "name":"PUNJAB_SLDC",
     "urls":["https://punjabsldc.org/","http://www.punjabsldc.org/","https://pstcl.org/"]},
    {"state":"Rajasthan",       "region":"Northern",      "name":"RAJASTHAN_SLDC",
     "urls":["https://rajsldc.com/","http://rajsldc.com/","https://rvpn.co.in/"]},
    {"state":"Uttar Pradesh",   "region":"Northern",      "name":"UP_SLDC",
     "urls":["https://upsldc.org/","http://www.upsldc.org/","https://upptcl.org/"]},
    {"state":"Uttarakhand",     "region":"Northern",      "name":"UK_SLDC",
     "urls":["https://ptcul.org/","https://www.ptcul.org/sldc"]},
    # Western Region
    {"state":"Gujarat",         "region":"Western",       "name":"GUJARAT_SLDC",
     "urls":["https://sldcguj.com/","http://www.sldcguj.com/"]},
    {"state":"Maharashtra",     "region":"Western",       "name":"MAHA_SLDC",
     "urls":["https://mahasldc.in/","http://www.mahasldc.in/","https://mahatransco.in/"]},
    {"state":"Madhya Pradesh",  "region":"Western",       "name":"MP_SLDC",
     "urls":["https://sldcmpindia.in/","http://www.sldcmpindia.com/"]},
    {"state":"Chhattisgarh",    "region":"Western",       "name":"CG_SLDC",
     "urls":["https://sldccg.com/","http://www.sldccg.gov.in/"]},
    {"state":"Goa",             "region":"Western",       "name":"GOA_SLDC",
     "urls":["https://goaelectricity.gov.in/","https://www.goaelectricity.gov.in/"]},
    # Southern Region
    {"state":"Andhra Pradesh",  "region":"Southern",      "name":"AP_SLDC",
     "urls":["https://apsldc.in/","https://sldc.aptransco.co.in/","https://aptransco.co.in/"]},
    {"state":"Karnataka",       "region":"Southern",      "name":"KA_SLDC",
     "urls":["https://kptclsldc.in/","https://www.kptcl.com/sldc","https://kptcl.com/"]},
    {"state":"Kerala",          "region":"Southern",      "name":"KERALA_SLDC",
     "urls":["https://sldckerala.com/","http://www.sldckerala.com/","https://kseb.in/"]},
    {"state":"Tamil Nadu",      "region":"Southern",      "name":"TN_SLDC",
     "urls":["https://www.tnebltd.gov.in/","https://sldc.tnebltd.gov.in/","https://tantransco.tn.gov.in/"]},
    {"state":"Telangana",       "region":"Southern",      "name":"TS_SLDC",
     "urls":["https://tssldc.in/","https://tstransco.cgg.gov.in/","https://tsspdcl.co.in/"]},
    {"state":"Puducherry",      "region":"Southern",      "name":"PUDUCHERRY_SLDC",
     "urls":["https://electricity.py.gov.in/"]},
    # Eastern Region
    {"state":"West Bengal",     "region":"Eastern",       "name":"WB_SLDC",
     "urls":["https://wbsedcl.in/","https://wbsetcl.in/","https://wbseb.gov.in/"]},
    {"state":"Odisha",          "region":"Eastern",       "name":"ODISHA_SLDC",
     "urls":["https://sldcorissa.org.in/","http://www.sldcorissa.org.in/"]},
    {"state":"Bihar",           "region":"Eastern",       "name":"BIHAR_SLDC",
     "urls":["https://bsptcl.bih.nic.in/","https://bsphcl.bih.nic.in/","https://nbpdcl.co.in/"]},
    {"state":"Jharkhand",       "region":"Eastern",       "name":"JH_SLDC",
     "urls":["https://jbvnl.co.in/","https://www.jbvnl.co.in/","https://jseb.jharkhand.gov.in/"]},
    {"state":"Sikkim",          "region":"Eastern",       "name":"SIKKIM_SLDC",
     "urls":["https://energy.sikkim.gov.in/","https://www.sikkim.gov.in/departments/energy"]},
    # North-Eastern Region
    {"state":"Assam",           "region":"North-Eastern", "name":"ASSAM_SLDC",
     "urls":["https://aegcl.co.in/","https://www.aegcl.co.in/","http://www.aegclsldc.org/"]},
    {"state":"Meghalaya",       "region":"North-Eastern", "name":"MEGHALAYA_SLDC",
     "urls":["https://meeclsldc.nic.in/","https://meeacl.gov.in/","http://meeclsldc.nic.in/"]},
    {"state":"Mizoram",         "region":"North-Eastern", "name":"MIZORAM_SLDC",
     "urls":["http://sldc.mizoram.gov.in/","https://powermizo.gov.in/"]},
    {"state":"Manipur",         "region":"North-Eastern", "name":"MANIPUR_SLDC",
     "urls":["https://manipurelectricity.gov.in/","https://mspcl.gov.in/"]},
    {"state":"Nagaland",        "region":"North-Eastern", "name":"NAGALAND_SLDC",
     "urls":["https://dpu.nagaland.gov.in/","https://nagalandpower.nic.in/"]},
    {"state":"Tripura",         "region":"North-Eastern", "name":"TRIPURA_SLDC",
     "urls":["https://tsecl.nic.in/","https://www.tsecl.in/"]},
    {"state":"Arunachal Pradesh","region":"North-Eastern","name":"AR_SLDC",
     "urls":["https://aphydro.org/","https://arunachalpower.nic.in/"]},
]

# Patterns to extract key data from SLDC pages
SLDC_PATTERNS = {
    "demand":      [
        r"(?:demand|drawl|drawal|load met|requirement)[^0-9]{0,50}([\d,]+(?:\.\d+)?)\s*(?:mw|mu|mwh)",
        r"([\d,]+(?:\.\d+)?)\s*(?:mw|mu)[^a-z]{0,20}(?:demand|drawl|load)",
        r"(?:peak demand|pd)[^0-9]{0,30}:\s*([\d,]+)",
    ],
    "generation":  [
        r"(?:generation|injection|infeed|gen\.?)[^0-9]{0,50}([\d,]+(?:\.\d+)?)\s*(?:mw|mu|mwh)",
        r"(?:total gen|tg)[^0-9]{0,20}:\s*([\d,]+)",
    ],
    "peak_demand": [
        r"(?:all.time peak|peak demand in fy)[^0-9]{0,50}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)",
        r"(?:peak demand)\s*:\s*([\d,]+(?:\.\d+)?)\s*(?:mw)",
        r"(?:maximum demand)[^0-9]{0,30}([\d,]+)",
    ],
    "freq":        [
        r"(?:frequency|freq\.?)[^0-9]{0,20}(4[89]\.\d{2,3}|50\.\d{2,3})\s*(?:hz)?",
        r"(4[89]\.\d{2,3}|50\.\d{2,3})\s*(?:hz|Hz)",
    ],
    "deficit":     [
        r"(?:deficit|shortage|shortfall)[^0-9]{0,40}([\d,]+(?:\.\d+)?)\s*(?:mw|mu)",
    ],
    "energy_today":[
        r"(?:energy (?:met|consumed|generated)|today.s energy)[^0-9]{0,40}([\d,]+(?:\.\d+)?)\s*(?:mu|mwh|gwh)",
    ],
}

def _extract_sldc_values(text):
    """Extract key metrics from SLDC page text using regex patterns."""
    vals = {}
    for key, patterns in SLDC_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                g = m.group(1) if key != "freq" else m.group(0)
                v = _to_float(g)
                if v is None: continue
                if key == "freq":
                    if 48 < v < 52: vals[key] = v; break
                else:
                    if v > 0: vals[key] = v; break
    return vals

def scrape_one_sldc(sldc_info):
    """Scrape a single state SLDC website."""
    name   = sldc_info["name"]
    state  = sldc_info["state"]
    region = sldc_info["region"]
    result = {"status":"error","name":name,"state":state,"region":region,"ts":ts()}

    html = None
    for url in sldc_info["urls"]:
        # Try Playwright first
        try:
            html = pw_get(url)
            if html and len(html) > 1000: break
        except: pass
        # Plain requests fallback
        try:
            resp = get_html(url, referer=url)
            if resp and len(resp.text) > 500: html = resp.text; break
        except: pass

    if not html:
        log_scrape(name, "unreachable", "All URLs failed")
        result["error"] = "Unreachable"
        return result

    soup = BeautifulSoup(html, "lxml")
    full_text = soup.get_text(" ")
    vals = _extract_sldc_values(full_text)

    # For JS-rendered SLDCs: try Playwright element-level extraction
    # targeting elements that likely contain demand/frequency data
    if not any(vals.values()):
        for url in sldc_info["urls"]:
            try:
                from playwright.sync_api import sync_playwright
                with sync_playwright() as pw:
                    browser = pw.chromium.launch(headless=True, args=["--no-sandbox"])
                    page = browser.new_page(user_agent=UA)
                    page.goto(url, timeout=30000)
                    # Wait for any element matching demand/frequency keywords
                    for sel in ["[class*=demand]","[class*=mw]","[class*=freq]","[class*=load]",
                                "[id*=demand]","[id*=freq]","table"]:
                        try: page.wait_for_selector(sel, timeout=5000); break
                        except: continue
                    # Extract all visible text from matching elements
                    for sel in ["[class*=demand]","[class*=mw]","[class*=load]","[class*=freq]",
                                "[id*=demand]","[id*=freq]","table","body"]:
                        try:
                            elems = page.query_selector_all(sel)
                            combined = " ".join(e.inner_text() for e in elems[:20] if e.is_visible())
                            v2 = _extract_sldc_values(combined)
                            if any(v2.values()):
                                vals.update({k:v for k,v in v2.items() if v and not vals.get(k)})
                        except: pass
                    browser.close()
                if any(vals.values()): break
            except ImportError: break
            except Exception as e:
                print(f"  SLDC PW element extract ({name}): {e}", file=sys.stderr)
                break

    # Supplement with table parsing — many SLDCs show live data in HTML tables
    if not vals.get("demand"):
        for tbl in soup.find_all("table"):
            rows = tbl.find_all("tr")
            for row in rows:
                cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
                row_text = " ".join(cells).lower()
                if any(k in row_text for k in ["demand","drawl","generation","frequency"]):
                    for i, cell in enumerate(cells):
                        v = _to_float(cell)
                        if v and v > 100:  # MW values
                            if "demand" in row_text and not vals.get("demand"): vals["demand"] = v
                            if "gen" in row_text and not vals.get("generation"): vals["generation"] = v
                        if v and 48 < v < 52 and not vals.get("freq"): vals["freq"] = v

    # Peak demand + date
    peak_date = None
    for tag in soup.find_all(string=lambda t: t and "peak" in t.lower() and "demand" in t.lower()):
        parent = tag.parent
        if parent:
            date_m = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2})", parent.get_text())
            if date_m: peak_date = date_m.group(1)

    now = ts()
    con = db()
    con.execute("""INSERT INTO sldc_data
        (scraped_at,sldc,state,region,freq_hz,demand_mw,generation_mw,
         peak_demand_mw,peak_demand_date,energy_today_mu,deficit_mw,surplus_mw,raw_snippet)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (now, name, state, region,
         vals.get("freq"), vals.get("demand"), vals.get("generation"),
         vals.get("peak_demand"), peak_date,
         vals.get("energy_today"), vals.get("deficit"), None,
         full_text[:500]))
    con.commit(); con.close()

    has_data = any(vals.values())
    status = "ok" if has_data else "reached_no_data"
    log_scrape(name, status, rows=1 if has_data else 0)
    print(f"  {'✓' if has_data else '~'} {name}: demand={vals.get('demand')}, gen={vals.get('generation')}, freq={vals.get('freq')}", file=sys.stderr)
    result.update({"status":"ok","state":state,"region":region,
                   "demand_mw":vals.get("demand"),"generation_mw":vals.get("generation"),
                   "freq_hz":vals.get("freq"),"peak_demand_mw":vals.get("peak_demand"),
                   "deficit_mw":vals.get("deficit"),"energy_today_mu":vals.get("energy_today"),
                   "ts":now})
    return result

def scrape_all_sldcs():
    """Scrape all SLDC websites in parallel."""
    import concurrent.futures
    results = {}
    # Run up to 5 SLDCs in parallel to speed things up
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(scrape_one_sldc, s): s for s in SLDCS}
        for fut in concurrent.futures.as_completed(futures):
            sldc = futures[fut]
            try:
                results[sldc["name"]] = fut.result()
            except Exception as e:
                log_scrape(sldc["name"], "error", str(e))
                results[sldc["name"]] = {"status":"error","error":str(e),"state":sldc["state"]}
    return results

def get_sldc_latest():
    """Get latest SLDC reading per state."""
    return query("""SELECT s.* FROM sldc_data s
        INNER JOIN (SELECT sldc, MAX(scraped_at) max_ts FROM sldc_data GROUP BY sldc) m
        ON s.sldc=m.sldc AND s.scraped_at=m.max_ts ORDER BY s.state""")

# ════════════════════════════════════════════════════════════════════════════
# 10b. GOA STATE — Dedicated granular scraper
#      Sources:
#        Goa Electricity Dept / SLDC: https://goaelectricity.gov.in/
#        WRLDC (covers Goa):          https://wrldc.in/
#        Vidyut PRAVAH (Goa state):   already in scrape_vidyut_pravah()
#        CEA monthly state data:      already in scrape_cea_state_data()
#        POSOCO/Grid-India:           covers Goa in WR aggregates
# ════════════════════════════════════════════════════════════════════════════

# Goa's major power purchase sources (from GPDCL/GED annual reports & PPA data)
GOA_POWER_SOURCES = [
    # Central sector allocations (Goa's share)
    {"source": "NTPC Mouda",        "type": "Thermal",    "capacity_mw": 180,  "agency": "Central"},
    {"source": "NTPC Korba",        "type": "Thermal",    "capacity_mw": 92,   "agency": "Central"},
    {"source": "NTPC Sipat",        "type": "Thermal",    "capacity_mw": 120,  "agency": "Central"},
    {"source": "NTPC Gas (Kayamkulam/Gandhar)", "type": "Gas", "capacity_mw": 25, "agency": "Central"},
    {"source": "Tata Power (Trombay)", "type": "Thermal", "capacity_mw": 48,   "agency": "IPP"},
    {"source": "NPC (Tarapur)",     "type": "Nuclear",    "capacity_mw": 25,   "agency": "Central"},
    {"source": "Sardar Sarovar",    "type": "Hydro",      "capacity_mw": 30,   "agency": "Central"},
    {"source": "Koyna (MSEB)",      "type": "Hydro",      "capacity_mw": 20,   "agency": "State"},
    {"source": "Indira Sagar",      "type": "Hydro",      "capacity_mw": 18,   "agency": "Central"},
    {"source": "Salal / Uri / Chamera", "type": "Hydro",  "capacity_mw": 22,   "agency": "Central"},
    # Local / State sector
    {"source": "Goa Solar (GEDCL Rooftop)", "type": "Solar", "capacity_mw": 85, "agency": "State"},
    {"source": "Goa Wind",          "type": "Wind",       "capacity_mw": 5,    "agency": "State"},
    {"source": "Waste to Energy (Saligao)", "type": "WtE", "capacity_mw": 8,   "agency": "State"},
    # Short-term / Exchange purchases
    {"source": "IEX DAM/RTM",       "type": "Exchange",   "capacity_mw": None, "agency": "Exchange"},
    {"source": "Bilateral (Short-term)", "type": "Bilateral", "capacity_mw": None, "agency": "Market"},
]

GOA_SLDC_URLS = [
    "https://goaelectricity.gov.in/",
    "https://www.goaelectricity.gov.in/",
    "https://goaelectricity.gov.in/sldc/",
    "https://goaelectricity.gov.in/real-time/",
]

def scrape_goa_state():
    """
    Scrape comprehensive Goa state power data:
    - Real-time demand/supply/frequency from Goa SLDC
    - Power purchase source breakdown from WRLDC state drawal
    - Historical trend from stored DB
    Returns enriched dict for the dedicated Goa tab.
    """
    result = {
        "status": "partial",
        "state": "Goa",
        "region": "Western",
        "ts": ts(),
        "sources": GOA_POWER_SOURCES,
    }

    # ── 1. Try Goa SLDC website ────────────────────────────────────────────
    sldc_result = scrape_one_sldc({
        "state": "Goa", "region": "Western", "name": "GOA_SLDC",
        "urls": GOA_SLDC_URLS,
    })
    if sldc_result.get("demand_mw"):
        result.update({
            "demand_mw":       sldc_result.get("demand_mw"),
            "generation_mw":   sldc_result.get("generation_mw"),
            "freq_hz":         sldc_result.get("freq_hz"),
            "peak_demand_mw":  sldc_result.get("peak_demand_mw"),
            "energy_today_mu": sldc_result.get("energy_today_mu"),
            "deficit_mw":      sldc_result.get("deficit_mw"),
            "status":          "ok",
        })

    # ── 2. Pull Goa rows from existing DB tables ───────────────────────────
    con = db()
    try:
        # Latest from sldc_data
        sldc_rows = con.execute(
            "SELECT * FROM sldc_data WHERE state='Goa' ORDER BY scraped_at DESC LIMIT 5"
        ).fetchall()
        if sldc_rows:
            cols = [d[0] for d in con.execute("PRAGMA table_info(sldc_data)").fetchall()]
            latest = dict(zip(cols, sldc_rows[0]))
            result.setdefault("demand_mw",       latest.get("demand_mw"))
            result.setdefault("generation_mw",   latest.get("generation_mw"))
            result.setdefault("freq_hz",         latest.get("freq_hz"))
            result.setdefault("peak_demand_mw",  latest.get("peak_demand_mw"))
            result.setdefault("energy_today_mu", latest.get("energy_today_mu"))
            result.setdefault("deficit_mw",      latest.get("deficit_mw"))

        # Historical peak demand (last 90 days)
        peak_rows = con.execute(
            """SELECT date, peak_demand_mw FROM peak_demand
               WHERE entity_name='Goa' AND date >= date('now','-90 days')
               ORDER BY date"""
        ).fetchall()
        result["peak_trend"] = [{"date": r[0], "peak_mw": r[1]} for r in peak_rows]

        # Historical demand trend from sldc_data (last 30 days)
        trend_rows = con.execute(
            """SELECT date(scraped_at) day,
                      AVG(demand_mw) avg_dem, MAX(demand_mw) max_dem,
                      AVG(freq_hz) avg_freq
               FROM sldc_data WHERE state='Goa'
                 AND scraped_at >= datetime('now','-30 days')
               GROUP BY date(scraped_at) ORDER BY day"""
        ).fetchall()
        result["demand_trend"] = [
            {"date": r[0], "avg_mw": r[1], "peak_mw": r[2], "freq": r[3]}
            for r in trend_rows
        ]

        # State demand from cea/vidyut pravah (energy + peak)
        state_rows = con.execute(
            """SELECT date, energy_req_mu, energy_avail_mu, energy_deficit_mu,
                      peak_demand_mw, peak_met_mw, source
               FROM state_demand WHERE state='Goa'
               ORDER BY scraped_at DESC LIMIT 30"""
        ).fetchall()
        result["state_supply_history"] = [
            {"date": r[0], "req_mu": r[1], "avail_mu": r[2], "deficit_mu": r[3],
             "peak_dem_mw": r[4], "peak_met_mw": r[5], "source": r[6]}
            for r in state_rows
        ]

        # WRLDC state drawal for Goa
        wrldc_rows = con.execute(
            """SELECT scraped_at, scheduled_mw, actual_mw, deviation_mw, freq_hz
               FROM rldc_state_drawal WHERE LOWER(state) LIKE '%goa%'
               ORDER BY scraped_at DESC LIMIT 48"""
        ).fetchall()
        result["wrldc_drawal"] = [
            {"ts": r[0], "scheduled_mw": r[1], "actual_mw": r[2],
             "deviation_mw": r[3], "freq_hz": r[4]}
            for r in wrldc_rows
        ]

        result["status"] = "ok"
    except Exception as e:
        print(f"  Goa DB query error: {e}", file=sys.stderr)
    finally:
        con.close()

    log_scrape("GOA_STATE", result["status"], rows=1)
    return result

def get_goa_data():
    """Return Goa state data bundle for the dashboard API."""
    return scrape_goa_state()

def get_goa_history(days=30):
    """Historical Goa data for trend charts."""
    cutoff = (datetime.date.today() - datetime.timedelta(days=days)).isoformat()
    return {
        "sldc_trend": query(
            """SELECT date(scraped_at) day, AVG(demand_mw) avg_dem,
                      MAX(demand_mw) max_dem, AVG(freq_hz) avg_freq,
                      AVG(generation_mw) avg_gen
               FROM sldc_data WHERE state='Goa' AND scraped_at>=?
               GROUP BY date(scraped_at) ORDER BY day""",
            (cutoff + " 00:00:00",)
        ),
        "peak_trend": query(
            "SELECT date, peak_demand_mw FROM peak_demand WHERE entity_name='Goa' AND date>=? ORDER BY date",
            (cutoff,)
        ),
        "state_demand": query(
            """SELECT date, energy_req_mu, energy_avail_mu, energy_deficit_mu,
                      peak_demand_mw, peak_met_mw
               FROM state_demand WHERE state='Goa' AND date>=? ORDER BY date""",
            (cutoff,)
        ),
    }



# ════════════════════════════════════════════════════════════════════════════
# 11. NSE + MCX Electricity Futures  (launched July 2025)
# ════════════════════════════════════════════════════════════════════════════
NSE_HEADERS = {
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/",
    "X-Requested-With": "XMLHttpRequest",
    "Connection": "keep-alive",
}

# NSE commodity derivatives API — ELECMBL
# NSE requires: GET homepage first (sets cookies), then call API with those cookies
NSE_API_FUTURES = [
    "https://www.nseindia.com/api/quote-commodity?symbol=ELECMBL",
    "https://www.nseindia.com/api/allcontracts?type=COMMODITY&symbol=ELECMBL",
    "https://www.nseindia.com/api/live-analysis-quote-equities?symbol=ELECMBL",
    "https://www.nseindia.com/api/quote-derivative?symbol=ELECMBL",
    "https://www.nseindia.com/api/marketStatus",  # fallback: at least get status
]
NSE_COOKIE_URL = "https://www.nseindia.com/"
NSE_FUTURES_PAGE = "https://www.nseindia.com/static/products-services/electricity-futures"
NSE_MARKET_DATA_URL = "https://www.nseindia.com/market-data/live-market-commodity-derivatives"

MCX_FUTURES_URLS = [
    "https://www.mcxindia.com/market-data/live-market/commodity-futures",
    "https://www.mcxindia.com/market-data/commodity-market-data",
    "https://www.mcxindia.com/",
]
MCX_API_URLS = [
    "https://www.mcxindia.com/api/sitemap/GetTickerData",
    "https://www.mcxindia.com/api/LiveMarket/GetLiveMarketData",
]

def _get_nse_session():
    """NSE requires a valid session cookie before API calls work."""
    session = requests.Session()
    session.headers.update(NSE_HEADERS)
    try:
        session.get(NSE_COOKIE_URL, timeout=15, verify=False)
    except: pass
    return session

def scrape_nse_electricity_futures():
    """Scrape NSE Monthly Electricity Futures (ELECMBL) data."""
    result = {"status":"error","exchange":"NSE","contracts":[],"ts":ts()}

    # Method 1: NSE JSON API (session required)
    session = _get_nse_session()
    for api_url in NSE_API_FUTURES:
        try:
            r = session.get(api_url, timeout=20, verify=False)
            if r.ok and "json" in r.headers.get("Content-Type",""):
                data = r.json()
                contracts = _parse_nse_elec_futures(data)
                if contracts:
                    _save_futures(contracts, "NSE")
                    log_scrape("NSE_ELECMBL","ok",rows=len(contracts))
                    print(f"  ✓ NSE Electricity Futures (API): {len(contracts)} contracts", file=sys.stderr)
                    result.update({"status":"ok","contracts":contracts,"ts":ts()})
                    return result
        except Exception as e:
            print(f"  ✗ NSE API ({api_url}): {e}", file=sys.stderr)

    # Method 2: Playwright on NSE futures page + market data page
    for nse_pw_url in [NSE_FUTURES_PAGE, NSE_MARKET_DATA_URL]:
      try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            # --disable-http2 bypasses NSE HTTP/2 fingerprint blocking
            browser = pw.chromium.launch(headless=True, args=["--disable-http2","--no-sandbox"])
            page = browser.new_page(user_agent=UA)
            # Set cookies from homepage first
            page.goto(NSE_COOKIE_URL, timeout=20000)
            import time as _t; _t.sleep(1)
            page.goto(nse_pw_url, timeout=30000)
            try: page.wait_for_selector("table, .trading-data, [class*=futures]", timeout=10000)
            except: pass
            page.wait_for_load_state("networkidle", timeout=10000)
            html = page.content()
            browser.close()
        soup = BeautifulSoup(html, "lxml")
        contracts = _parse_futures_html(soup, "NSE")
        if contracts:
            _save_futures(contracts, "NSE")
            log_scrape("NSE_ELECMBL","ok",rows=len(contracts))
            print(f"  ✓ NSE Electricity Futures (PW/{nse_pw_url}): {len(contracts)} contracts", file=sys.stderr)
            result.update({"status":"ok","contracts":contracts,"ts":ts()})
            return result
        # Even if no structured contracts, extract any price data from text
        full = soup.get_text(" ")
        pm = re.findall(r"(?:4[0-9]{3}|5[0-9]{3})\.\d*", full)
        if pm: print(f"  ~ NSE price mentions: {pm[:3]}", file=sys.stderr)
      except Exception as e:
        print(f"  ✗ NSE Playwright ({nse_pw_url}): {e}", file=sys.stderr)

    # Method 3: End-of-day bhavcopy CSV (NSE publishes daily)
    try:
        today_dt  = datetime.date.today()
        today_str = today_dt.strftime("%d%b%Y").upper()
        ddmmyyyy  = today_dt.strftime("%d%m%Y")
        # Try multiple bhavcopy URL formats NSE uses
        for bhav_url in [
            f"https://nsearchives.nseindia.com/content/fo/BhavCopy_NSE_FO_0_0_0_{today_str}_F_0000.csv",
            f"https://nsearchives.nseindia.com/archives/fo/bhavcopy/fobhav{ddmmyyyy}.csv",
            f"https://nsearchives.nseindia.com/content/com/BhavCopy_NSE_COM_0_0_0_{today_str}_F_0000.csv",
        ]:
            try:
                r2 = session.get(bhav_url, timeout=20, verify=False)
                if not r2.ok: continue
                df = pd.read_csv(io.StringIO(r2.text))
                elec = df[df.apply(lambda row: any("ELEC" in str(v).upper() for v in row), axis=1)]
                if elec.empty: continue
                contracts = []
                for _, row in elec.iterrows():
                    contracts.append({
                        "symbol": str(row.get("SYMBOL","ELECMBL")),
                        "expiry_date": str(row.get("EXPIRY_DT","")),
                        "last_price": _to_float(row.get("CLOSE_PRICE") or row.get("LAST_PRICE")),
                        "open_price": _to_float(row.get("OPEN_PRICE")),
                        "high_price": _to_float(row.get("HIGH_PRICE")),
                        "low_price": _to_float(row.get("LOW_PRICE")),
                        "volume_lots": _to_float(row.get("TTL_TRD_QNTY") or row.get("VOLUME")),
                        "exchange": "NSE"
                    })
                if contracts:
                    _save_futures(contracts, "NSE")
                    log_scrape("NSE_ELECMBL","ok",rows=len(contracts))
                    result.update({"status":"ok","contracts":contracts,"ts":ts()})
                    return result
            except: continue
    except Exception as e:
        print(f"  ✗ NSE bhavcopy: {e}", file=sys.stderr)


    log_scrape("NSE_ELECMBL","error","All NSE methods failed")
    result["error"] = "NSE: session/cookie based API — may need manual cookie injection"
    return result

def _parse_nse_elec_futures(data):
    """Parse NSE API response for electricity futures contracts."""
    contracts = []
    # NSE typically returns: {"records":{"data":[...]}} or {"data":[...]}
    items = []
    if isinstance(data, dict):
        items = (data.get("records",{}).get("data") or
                 data.get("data") or
                 data.get("FutChainData") or [])
    elif isinstance(data, list):
        items = data
    for item in (items if isinstance(items, list) else []):
        if not isinstance(item, dict): continue
        sym = str(item.get("identifier","") or item.get("symbol",""))
        if "ELEC" not in sym.upper() and "ELECTRICITY" not in sym.upper(): continue
        contracts.append({
            "symbol":        sym,
            "expiry_date":   str(item.get("expiryDate","") or item.get("expiry","")),
            "last_price":    _to_float(item.get("lastPrice") or item.get("last")),
            "open_price":    _to_float(item.get("open") or item.get("openPrice")),
            "high_price":    _to_float(item.get("high") or item.get("highPrice")),
            "low_price":     _to_float(item.get("low") or item.get("lowPrice")),
            "close_price":   _to_float(item.get("close") or item.get("closePrice")),
            "prev_close":    _to_float(item.get("previousClose") or item.get("prevClose")),
            "change_rs":     _to_float(item.get("change") or item.get("pChange")),
            "change_pct":    _to_float(item.get("pChange") or item.get("changePct")),
            "volume_lots":   _to_float(item.get("numberOfContractsTraded") or item.get("volume") or item.get("tradedVolume")),
            "open_interest": _to_float(item.get("openInterest") or item.get("oi")),
            "turnover_cr":   _to_float(item.get("totalTurnover") or item.get("turnover")),
            "vwap":          _to_float(item.get("vwap") or item.get("weightedAverage")),
            "exchange":      "NSE",
            "unit":          "₹/MWh"
        })
    return contracts

def _parse_futures_html(soup, exchange):
    """Parse futures data from HTML table."""
    contracts = []
    for tbl in soup.find_all("table"):
        rows = tbl.find_all("tr")
        if len(rows) < 2: continue
        hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["th","td"])]
        if not any(k in " ".join(hdrs) for k in ["price","ltp","last","expiry","contract"]): continue
        col = {}
        for i,h in enumerate(hdrs):
            if "expiry" in h:    col.setdefault("expiry",i)
            if "last" in h or "ltp" in h or "close" in h: col.setdefault("last",i)
            if "open" in h:      col.setdefault("open",i)
            if "high" in h:      col.setdefault("high",i)
            if "low" in h:       col.setdefault("low",i)
            if "volume" in h or "qty" in h: col.setdefault("volume",i)
            if "oi" in h or "interest" in h: col.setdefault("oi",i)
        for row in rows[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(cells) < 2: continue
            lp = _to_float(cells[col.get("last",1)])
            if lp and lp > 1000:  # electricity futures typically >₹1000/MWh
                contracts.append({
                    "symbol":      "ELECMBL",
                    "expiry_date": cells[col.get("expiry",0)] if "expiry" in col else "",
                    "last_price":  lp,
                    "open_price":  _to_float(cells[col.get("open",2)]) if "open" in col else None,
                    "high_price":  _to_float(cells[col.get("high",3)]) if "high" in col else None,
                    "low_price":   _to_float(cells[col.get("low",4)]) if "low" in col else None,
                    "volume_lots": _to_float(cells[col.get("volume",5)]) if "volume" in col else None,
                    "open_interest":_to_float(cells[col.get("oi",6)]) if "oi" in col else None,
                    "exchange":    exchange,
                    "unit":        "₹/MWh"
                })
        if contracts: break
    return contracts

def scrape_mcx_electricity_futures():
    """Scrape MCX Electricity Futures (launched July 10, 2025)."""
    result = {"status":"error","exchange":"MCX","contracts":[],"ts":ts()}

    # Try MCX JSON API first
    for api_url in MCX_API_URLS:
        try:
            h = dict(HEADERS); h["Referer"]="https://www.mcxindia.com/"; h["X-Requested-With"]="XMLHttpRequest"
            r = requests.get(api_url, headers=h, timeout=20, verify=False)
            if r.ok and "json" in r.headers.get("Content-Type",""):
                data_j = r.json()
                items = data_j if isinstance(data_j,list) else data_j.get("data",data_j.get("Data",[]))
                contracts = []
                for item in (items if isinstance(items,list) else []):
                    sym = str(item.get("Symbol","") or item.get("symbol",""))
                    if "ELEC" not in sym.upper() and "POWER" not in sym.upper(): continue
                    lp = _to_float(item.get("LastTradedPrice") or item.get("ltp") or item.get("Close"))
                    if lp: contracts.append({"symbol":sym,"last_price":lp,
                        "open_price":_to_float(item.get("Open")),
                        "high_price":_to_float(item.get("High")),
                        "low_price":_to_float(item.get("Low")),
                        "volume_lots":_to_float(item.get("Volume") or item.get("TradedQty")),
                        "expiry_date":str(item.get("ExpiryDate","")),
                        "exchange":"MCX","unit":"₹/MWh"})
                if contracts:
                    _save_futures(contracts,"MCX"); log_scrape("MCX_ELEC","ok",rows=len(contracts))
                    print(f"  ✓ MCX (API): {len(contracts)} contracts", file=sys.stderr)
                    result.update({"status":"ok","contracts":contracts,"ts":ts()}); return result
        except Exception as e: print(f"  ✗ MCX API ({api_url}): {e}", file=sys.stderr)

    for url in MCX_FUTURES_URLS:
        try:
            html = None
            try: html = pw_get(url)
            except: pass
            if not html:
                resp = get_html(url, referer="https://www.mcxindia.com/")
                html = resp.text
            soup = BeautifulSoup(html, "lxml")

            # Try structured table parsing
            contracts = _parse_futures_html(soup, "MCX")

            # Fallback: scan text for electricity price mentions
            if not contracts:
                full = soup.get_text(" ")
                # MCX electricity futures ~ ₹4000-6000/MWh
                price_m = re.findall(r"(?:elec|power|electricity)[^0-9]{0,30}(4[0-9]{3}|5[0-9]{3}|6[0-9]{3})",
                                     full, re.IGNORECASE)
                if price_m:
                    contracts = [{"symbol":"MCX-ELEC","last_price":_to_float(price_m[0]),"exchange":"MCX","unit":"₹/MWh"}]

            if contracts:
                _save_futures(contracts, "MCX")
                log_scrape("MCX_ELEC","ok",rows=len(contracts))
                print(f"  ✓ MCX Electricity Futures: {len(contracts)} contracts", file=sys.stderr)
                result.update({"status":"ok","contracts":contracts,"ts":ts()})
                return result
        except Exception as e:
            print(f"  ✗ MCX ({url}): {e}", file=sys.stderr)

    log_scrape("MCX_ELEC","error","No MCX electricity futures data found")
    result["error"] = "MCX futures: no data found"
    return result

def _save_futures(contracts, exchange):
    now = ts(); con = db()
    for c in contracts:
        con.execute("""INSERT INTO electricity_futures
            (scraped_at,exchange,symbol,expiry_date,last_price,open_price,high_price,
             low_price,close_price,prev_close,change_rs,change_pct,volume_lots,
             open_interest,turnover_cr,vwap,unit)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (now, exchange, c.get("symbol",""), c.get("expiry_date",""),
             c.get("last_price"), c.get("open_price"), c.get("high_price"),
             c.get("low_price"), c.get("close_price"), c.get("prev_close"),
             c.get("change_rs"), c.get("change_pct"), c.get("volume_lots"),
             c.get("open_interest"), c.get("turnover_cr"), c.get("vwap"),
             c.get("unit","₹/MWh")))
    con.commit(); con.close()

def get_futures_latest():
    return query("""SELECT f.* FROM electricity_futures f
        INNER JOIN (SELECT exchange, symbol, MAX(scraped_at) max_ts
                    FROM electricity_futures GROUP BY exchange, symbol) m
        ON f.exchange=m.exchange AND f.symbol=m.symbol AND f.scraped_at=m.max_ts
        ORDER BY f.exchange, f.expiry_date""")

def get_futures_trend(days=30):
    c = (datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()
    return query("""SELECT date(scraped_at) day, exchange, symbol,
        AVG(last_price) avg_price, MIN(low_price) min_price, MAX(high_price) max_price,
        SUM(volume_lots) total_volume
        FROM electricity_futures WHERE scraped_at>? AND last_price IS NOT NULL
        GROUP BY day, exchange, symbol ORDER BY day""", (c,))

# ════════════════════════════════════════════════════════════════════════════
# MASTER SCRAPE
# ════════════════════════════════════════════════════════════════════════════
def scrape_all():
    print(f"\n[{ts()[:19]}] ══ Starting full scrape ══", file=sys.stderr)
    r = {
        "scraped_at":    ts(),
        "iex_dam":       scrape_iex_dam(),
        "pxil":          scrape_pxil(),
        "posoco":        scrape_posoco(),  # Grid-India PSP XLS + RT fallback
        "rldcs":         scrape_all_rldcs(),
        "rpcs":          scrape_all_rpcs(),
        "remc":          scrape_remc(),
        "vidyut_pravah": scrape_vidyut_pravah(),
        "cea_states":    scrape_cea_state_data(),
        "merit":         scrape_merit(),
        "sldcs":         scrape_all_sldcs(),
        "nse_futures":   scrape_nse_electricity_futures(),
        "mcx_futures":   scrape_mcx_electricity_futures(),
        "curtailment":   scrape_curtailment(),   # VRE curtailment from Grid India
        "akshayurja":    scrape_akshayurja(),    # RE installed capacity from MNRE
        "storage":       scrape_storage(),       # BESS + Pumped Hydro Storage
        "goa_state":     scrape_goa_state(),     # Dedicated Goa state scraper
    }
    ok     = [k for k in ["iex_dam","pxil","posoco","remc","vidyut_pravah","cea_states","merit","curtailment","akshayurja"] if r[k].get("status")=="ok"]
    rldc_ok= [k for k,v in r["rldcs"].items() if v.get("status")=="ok"]
    rpc_ok = [k for k,v in r["rpcs"].items()  if v.get("status")=="ok"]
    sldc_ok= [k for k,v in r["sldcs"].items() if v.get("status")=="ok"]

    # ── Record daily peak demand for national + all states every scrape ───
    _record_daily_peaks(r)

    print(f"[{ts()[:19]}] ══ Done: {len(ok)} main | {len(rldc_ok)}/5 RLDCs | {len(rpc_ok)}/5 RPCs | {len(sldc_ok)}/{len(SLDCS)} SLDCs ══\n", file=sys.stderr)
    return r


def _record_daily_peaks(scrape_result):
    """
    After every scrape, record current demand as a candidate daily peak.
    _save_peak_demand() uses MAX logic — only updates if this reading is higher
    than any previously recorded peak for today. Runs for:
      - National (All India) from POSOCO/Grid India
      - Per state from Vidyut PRAVAH + SLDC data
      - Per state from CEA monthly (peak_demand_mw field)
    """
    dt = today()
    try:
        # National peak — from POSOCO national demand
        pos = scrape_result.get("posoco", {})
        nat = pos.get("national", {})
        nat_mw = nat.get("demand_mw") or nat.get("peak_demand_mw")
        if nat_mw and nat_mw > 50000:
            _save_peak_demand(dt, nat_mw, "national", "All India")

        # Also capture from PSP XLS peak_demand_mw if available
        psp_peak = nat.get("peak_demand_mw")
        if psp_peak and psp_peak > 50000:
            _save_peak_demand(dt, psp_peak, "national", "All India")

        # State peaks — from Vidyut PRAVAH
        vp_states = scrape_result.get("vidyut_pravah", {}).get("states", [])
        for s in vp_states:
            mw = s.get("demand_mw") or s.get("peak_demand_mw")
            st = s.get("state")
            if st and mw and mw > 100:
                _save_peak_demand(dt, mw, "state", st)
                # If VP also has an explicit peak_demand_mw (some states report it)
                if s.get("peak_demand_mw") and s["peak_demand_mw"] > mw:
                    _save_peak_demand(dt, s["peak_demand_mw"], "state", st)

        # State peaks — from CEA monthly (these are monthly peaks, mark separately)
        cea_states = scrape_result.get("cea_states", {}).get("states", [])
        for s in cea_states:
            pm = s.get("peak_demand_mw")
            st = s.get("state")
            # CEA monthly peak is typically the month's peak — store as monthly entity_type
            if st and pm and pm > 100:
                con = db()
                month_str = datetime.date.today().strftime("%Y-%m")
                existing = con.execute(
                    "SELECT id, peak_demand_mw FROM peak_demand WHERE date=? AND entity_name=? AND entity_type='state_monthly'",
                    (month_str, st)).fetchone()
                if existing:
                    if pm > (existing[1] or 0):
                        con.execute("UPDATE peak_demand SET peak_demand_mw=?, updated_at=? WHERE id=?",
                                    (pm, ts(), existing[0]))
                else:
                    con.execute("INSERT INTO peak_demand (date,entity_type,entity_name,peak_demand_mw,peak_met_mw,updated_at) VALUES (?,?,?,?,?,?)",
                                (month_str, "state_monthly", st, pm, s.get("peak_met_mw"), ts()))
                con.commit(); con.close()

        # SLDC per-state peaks
        sldcs = scrape_result.get("sldcs", {})
        for sldc_name, sldc_r in sldcs.items():
            if not isinstance(sldc_r, dict): continue
            st  = sldc_r.get("state")
            mw  = sldc_r.get("demand_mw")
            pmw = sldc_r.get("peak_demand_mw")
            if st and pmw and pmw > 100:
                _save_peak_demand(dt, pmw, "state", st)
            elif st and mw and mw > 100:
                _save_peak_demand(dt, mw, "state", st)

    except Exception as e:
        print(f"  _record_daily_peaks error: {e}", file=sys.stderr)


def get_peak_demand_yearly(year=None):
    """
    Annual peak demand record per state + national.
    Returns: {state: {peak_mw, date, entity_type}, ...}
    """
    if year is None:
        year = datetime.date.today().year
    year_start = f"{year}-01-01"
    year_end   = f"{year}-12-31"

    # National annual peak
    nat_rows = query(
        """SELECT date, peak_demand_mw FROM peak_demand
           WHERE entity_type='national'
             AND date BETWEEN ? AND ?
             AND peak_demand_mw IS NOT NULL
           ORDER BY peak_demand_mw DESC LIMIT 1""",
        (year_start, year_end))

    # State annual peaks
    state_rows = query(
        """SELECT entity_name, MAX(peak_demand_mw) peak_mw,
                  date, entity_type
           FROM peak_demand
           WHERE entity_type IN ('state','state_monthly')
             AND date BETWEEN ? AND ?
             AND peak_demand_mw IS NOT NULL
           GROUP BY entity_name
           ORDER BY peak_mw DESC""",
        (year_start, year_end))

    return {
        "year":         year,
        "national":     nat_rows[0] if nat_rows else {},
        "states":       state_rows,
        "year_start":   year_start,
        "year_end":     year_end,
    }


def get_peak_yearly_for_dashboard(year=None):
    """Flat list of {state, peak_mw, date} for the dashboard _peakYearly field."""
    if year is None:
        year = datetime.date.today().year
    year_start = f"{year}-01-01"
    rows = query(
        """SELECT entity_name state, MAX(peak_demand_mw) peak_mw,
                  (SELECT date FROM peak_demand p2
                   WHERE p2.entity_name=p.entity_name
                     AND p2.date >= ?
                     AND p2.peak_demand_mw = MAX(p.peak_demand_mw)
                   LIMIT 1) date
           FROM peak_demand p
           WHERE entity_type IN ('state','state_monthly')
             AND date >= ?
             AND peak_demand_mw IS NOT NULL
           GROUP BY entity_name""",
        (year_start, year_start))
    return rows


# ── Query helpers ─────────────────────────────────────────────────────────────
def query(sql, params=()):
    con=db(); cur=con.execute(sql,params)
    cols=[d[0] for d in cur.description]; rows=[dict(zip(cols,r)) for r in cur.fetchall()]
    con.close(); return rows

def get_state_demand_latest():
    return query("SELECT s.* FROM state_demand s INNER JOIN (SELECT state, MAX(scraped_at) max_ts FROM state_demand GROUP BY state) m ON s.state=m.state AND s.scraped_at=m.max_ts ORDER BY COALESCE(s.shortage_pct,0) DESC")

def get_regional_latest():
    return query("SELECT r.* FROM regional_demand r INNER JOIN (SELECT region, MAX(scraped_at) max_ts FROM regional_demand GROUP BY region) m ON r.region=m.region AND r.scraped_at=m.max_ts")

def get_rldc_latest():
    return query("SELECT r.* FROM rldc_data r INNER JOIN (SELECT rldc, MAX(scraped_at) max_ts FROM rldc_data GROUP BY rldc) m ON r.rldc=m.rldc AND r.scraped_at=m.max_ts")

def get_peak_demand_national(days=30):
    """Daily national peak demand trend — 30-day history."""
    c = (datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()[:10]
    # Try peak_demand table — several possible entity_name values depending on scraper source
    rows = query(
        """SELECT date, peak_demand_mw FROM peak_demand
           WHERE (entity_type IN ('national','all_india','All India','ALL_INDIA')
                  OR entity_name IN ('ALL_INDIA','All India','National','all_india','INDIA'))
             AND date >= ? ORDER BY date""", (c,))
    if rows: return rows
    # Broader: any non-state, non-region row
    rows = query(
        """SELECT date, peak_demand_mw FROM peak_demand
           WHERE date>=? AND entity_type NOT IN ('state','region','rldc')
           ORDER BY date""", (c,))
    if rows: return rows
    # Fallback: derive daily MAX from national_demand table
    return query(
        """SELECT date(scraped_at) date, MAX(demand_mw) peak_demand_mw
           FROM national_demand
           WHERE scraped_at >= ? AND demand_mw IS NOT NULL AND demand_mw > 50000
           GROUP BY date(scraped_at) ORDER BY date""",
        (c + " 00:00:00",))

def get_peak_demand_states(days=30):
    """Daily peak demand per state."""
    c=(datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()[:10]
    rows=query("SELECT date,entity_name state,peak_demand_mw FROM peak_demand WHERE entity_type='state' AND date>=? ORDER BY date,state",(c,))
    if rows: return rows
    # Fallback: derive from state_demand
    return query("SELECT date(scraped_at) date, state, MAX(demand_mw) peak_demand_mw FROM state_demand WHERE scraped_at>=? AND demand_mw IS NOT NULL GROUP BY date(scraped_at),state ORDER BY date,state",(c+" 00:00:00",))

def get_national_trend(hours=48):
    """Return national demand readings for the last N hours.
    Default 48h but called with 720h (30 days) for the ratio chart,
    which needs multiple readings per day to compute daily averages."""
    c=(datetime.datetime.now()-datetime.timedelta(hours=hours)).isoformat()
    return query(
        "SELECT scraped_at, demand_mw, generation_mw, surplus_mw, deficit_mw, grid_freq_hz "
        "FROM national_demand WHERE scraped_at>? AND demand_mw IS NOT NULL "
        "ORDER BY scraped_at", (c,))

def get_state_trend(state,days=30):
    c=(datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()
    return query("SELECT * FROM state_demand WHERE state=? AND scraped_at>? ORDER BY scraped_at",(state,c))

def get_regional_trend(days=30):
    c=(datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()
    return query("SELECT region,date(scraped_at) day,AVG(demand_mw) avg_d,AVG(generation_mw) avg_g,AVG(interchange_mw) avg_i,AVG(deficit_mw) avg_def FROM regional_demand WHERE scraped_at>? GROUP BY region,day ORDER BY day",(c,))

def get_mcp_trend(days=30):
    c=(datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()
    return query("SELECT date,AVG(mcp) avg_mcp,MIN(mcp) min_mcp,MAX(mcp) max_mcp FROM iex_dam_summary WHERE scraped_at>? AND mcp IS NOT NULL GROUP BY date ORDER BY date",(c,))

def get_freq_trend(hours=24):
    c=(datetime.datetime.now()-datetime.timedelta(hours=hours)).isoformat()
    return query("SELECT scraped_at,grid_freq_hz hz FROM posoco_realtime WHERE scraped_at>? AND grid_freq_hz IS NOT NULL ORDER BY scraped_at",(c,))

def get_db_stats():
    con=db(); stats={}
    for t in ["iex_dam_summary","iex_dam_blocks","national_demand","regional_demand","rldc_data",
              "rldc_state_drawal","rpc_dsm_data","state_demand","merit_generation","remc_data",
              "posoco_realtime","pxil_prices","scrape_log"]:
        try:
            c,mn,mx=con.execute(f"SELECT COUNT(*),MIN(scraped_at),MAX(scraped_at) FROM {t}").fetchone()
            stats[t]={"count":c,"oldest":mn,"newest":mx}
        except: pass
    con.close(); return stats

def get_scrape_log(limit=50):
    return query("SELECT * FROM scrape_log ORDER BY scraped_at DESC LIMIT ?",(limit,))

def get_history(table, days=30, limit=500):
    safe={"iex_dam_summary","iex_dam_blocks","national_demand","regional_demand","rldc_data",
          "rldc_state_drawal","rpc_dsm_data","state_demand","merit_generation","remc_data",
          "posoco_realtime","pxil_prices","scrape_log"}
    if table not in safe: return []
    c=(datetime.datetime.now()-datetime.timedelta(days=days)).isoformat()
    return query(f"SELECT * FROM {table} WHERE scraped_at>? ORDER BY scraped_at DESC LIMIT ?",(c,limit))

# ── Excel export ──────────────────────────────────────────────────────────────
def export_excel(path=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    path=path or EXCEL_PATH; con=db()
    sheets={
        "National Demand":    "SELECT * FROM national_demand ORDER BY scraped_at DESC LIMIT 1000",
        "Regional Demand":    "SELECT * FROM regional_demand ORDER BY scraped_at DESC LIMIT 2000",
        "RLDC Data":          "SELECT * FROM rldc_data ORDER BY scraped_at DESC LIMIT 1000",
        "RLDC State Drawal":  "SELECT * FROM rldc_state_drawal ORDER BY scraped_at DESC LIMIT 3000",
        "RPC DSM Data":       "SELECT * FROM rpc_dsm_data ORDER BY scraped_at DESC LIMIT 2000",
        "State Demand":       "SELECT * FROM state_demand ORDER BY scraped_at DESC, shortage_pct DESC LIMIT 5000",
        "IEX DAM Summary":    "SELECT * FROM iex_dam_summary ORDER BY scraped_at DESC",
        "IEX DAM Blocks":     "SELECT * FROM iex_dam_blocks ORDER BY scraped_at DESC, block_no",
        "PXIL Prices":        "SELECT * FROM pxil_prices ORDER BY scraped_at DESC",
        "REMC Renewable":     "SELECT * FROM remc_data ORDER BY scraped_at DESC",
        "MERIT Generation":   "SELECT * FROM merit_generation ORDER BY scraped_at DESC, source",
        "POSOCO Realtime":    "SELECT * FROM posoco_realtime ORDER BY scraped_at DESC",
        "SLDC State Data":    "SELECT * FROM sldc_data ORDER BY scraped_at DESC, state LIMIT 5000",
        "Electricity Futures": "SELECT * FROM electricity_futures ORDER BY scraped_at DESC, exchange, expiry_date",
        "Peak Demand":        "SELECT * FROM peak_demand ORDER BY date DESC, entity_type, entity_name",
        "VRE Curtailment":    "SELECT * FROM curtailment_data ORDER BY date DESC",
        "RE Capacity":        "SELECT * FROM re_capacity ORDER BY scraped_at DESC",
        "Storage Dispatch":   "SELECT * FROM storage_dispatch ORDER BY scraped_at DESC",
        "Scrape Log":         "SELECT * FROM scrape_log ORDER BY scraped_at DESC",
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name,sql in sheets.items():
            try:
                df=pd.read_sql_query(sql,con); df.to_excel(writer,sheet_name=name,index=False)
                ws=writer.sheets[name]; fill=PatternFill("solid",start_color="0D1F38")
                for cell in ws[1]:
                    cell.font=Font(bold=True,color="00E5FF",size=10); cell.fill=fill
                    cell.alignment=Alignment(horizontal="center")
                ws.freeze_panes="A2"
                for col in ws.columns:
                    col_w=max((len(str(c.value or "")) for c in col),default=8)
                    ws.column_dimensions[col[0].column_letter].width=min(col_w+4,42)
            except Exception as e:
                print(f"  Skipping {name}: {e}",file=sys.stderr)
    con.close(); print(f"✓ Excel → {path}",file=sys.stderr); return path

# ── Demo data ─────────────────────────────────────────────────────────────────
def build_demo():
    import math, random as rnd
    now=datetime.datetime.now()
    blocks=[]
    for i in range(96):
        h,m=i//4,(i%4)*15; peak=(9<=h<=13)or(18<=h<=22); tr=2<=h<=5
        mcp=round(4+(1.3 if peak else -0.6 if tr else 0)+math.sin(i/8)*.35+rnd.uniform(-.25,.25),2)
        pw=int(5200+(3000 if peak else -1200 if tr else 0)+rnd.uniform(-300,300))
        t1=f"{h:02d}:{m:02d}"; h2=(h+(m+15)//60)%24; m2=(m+15)%60; t2=f"{h2:02d}:{m2:02d}"
        blocks.append({"block":str(i+1).zfill(2),"time":f"{t1}–{t2}","mcp":mcp,"purchase_mw":pw,"sell_mw":int(pw*rnd.uniform(.95,1.04))})
    bases={"Delhi":5800,"Maharashtra":21000,"Uttar Pradesh":22000,"Gujarat":16000,"Tamil Nadu":14000,"Rajasthan":12000,"West Bengal":9000,"Karnataka":11000,"Andhra Pradesh":10000,"Telangana":9500,"Madhya Pradesh":10500,"Punjab":7000,"Haryana":6000,"Bihar":6000,"Odisha":4500}
    state_data=[]
    for s in ALL_STATES:
        base=bases.get(s["name"],rnd.randint(800,4500))
        demand=int(base*rnd.uniform(.92,1.08)); sp=max(0,rnd.normalvariate(1.2,2.5)); met=int(demand*(1-sp/100))
        state_data.append({"state":s["name"],"region":s["region"],"demand_mw":demand,"met_mw":met,"shortage_mw":demand-met,"shortage_pct":round(sp,2),"energy_req_mu":round(demand*24/1000,1),"energy_avail_mu":round(met*24/1000,1),"energy_deficit_pct":round(sp,2),"peak_demand_mw":int(demand*1.08),"peak_met_mw":int(met*1.06)})
    regional={}
    for s in state_data:
        r=s["region"]
        if r not in regional: regional[r]={"region":r,"demand_mw":0,"generation_mw":0,"deficit_mw":0}
        regional[r]["demand_mw"]+=s["demand_mw"]; regional[r]["generation_mw"]+=s["met_mw"]+rnd.randint(-200,500); regional[r]["deficit_mw"]+=s["shortage_mw"]
    rldc_demo={}
    for reg,info in REGION_INFO.items():
        rd=regional.get(reg,{})
        rldc_demo[info["abbr"]+"LDC"]={"status":"ok","name":info["abbr"]+"LDC","region":reg,"freq":round(50+rnd.uniform(-.12,.12),3),"demand_mw":rd.get("demand_mw",0),"generation_mw":rd.get("generation_mw",0),"interchange_mw":rd.get("generation_mw",0)-rd.get("demand_mw",0),"deficit_mw":rd.get("deficit_mw",0),"state_drawal":[{"state":st,"scheduled_mw":int(bases.get(st,rnd.randint(500,5000))*rnd.uniform(.9,1.1)),"actual_mw":int(bases.get(st,rnd.randint(500,5000))*rnd.uniform(.88,1.12)),"deviation_mw":rnd.randint(-200,200)} for st in info["states"][:5]]}
    nd=sum(s["demand_mw"] for s in state_data); nm=sum(s["met_mw"] for s in state_data)
    freq_trend=[{"t":f"{(now-datetime.timedelta(minutes=(47-i)*5)).strftime('%H:%M')}","hz":round(50+rnd.uniform(-.15,.15),3)} for i in range(48)]
    rtm=[{"slot":f"{i>>1:02d}:{'30' if i%2 else '00'}","rtm_price":round(3.5+math.sin(i/6)*.7+rnd.uniform(-.3,.3),2)} for i in range(48)]
    return {
        "scraped_at":now.isoformat(),"is_demo":True,
        "iex_dam":{"status":"ok","mcp":4.28,"mcv_mwh":142300,"purchase_bid_mwh":158200,"sell_bid_mwh":176400,"blocks":blocks},
        "pxil":{"status":"ok","mcp":4.15,"mcv_mwh":8200,"exchange":"PXIL"},
        "posoco":{"status":"ok","national":{"freq":50.02,"demand_mw":nm+2000,"generation_mw":nm+4500,"surplus_mw":2500,"deficit_mw":nd-nm},"regional":list(regional.values())},
        "rldcs":rldc_demo,"rpcs":{"NRPC":{"status":"ok"},"WRPC":{"status":"ok"},"SRPC":{"status":"ok"},"ERPC":{"status":"ok"},"NERPC":{"status":"ok"}},
        "remc":{"status":"ok","solar_actual_mw":21500,"wind_actual_mw":12400,"solar_forecast_mw":22000,"wind_forecast_mw":12000},
        "vidyut_pravah":{"status":"ok","states":state_data},
        "cea_states":{"status":"ok","states":state_data},
        "merit":{"status":"ok","generation_table":[{"source":"Thermal (Coal)","generation_mw":108600},{"source":"Hydro","generation_mw":28700},{"source":"Nuclear","generation_mw":6800},{"source":"Solar","generation_mw":21500},{"source":"Wind","generation_mw":12400},{"source":"Gas","generation_mw":5100},{"source":"Other RES","generation_mw":3300}]},
        "_national":{"demand_mw":nm+2000,"generation_mw":nm+4500,"surplus_mw":2500,"deficit_mw":nd-nm,"freq":round(50+rnd.uniform(-.08,.08),3)},
        "_stateData":state_data,"_regionalData":list(regional.values()),
        "_rldcData":rldc_demo,
        "_freqTrend":freq_trend,"_rtmData":rtm,
        "_monthlyVolume":{"labels":["Dec","Jan","Feb","Mar","Apr","May"],"dam":[9140,9810,8320,10450,10920,11380],"rtm":[3210,3480,2980,3890,4120,4310]},
        "_nationalTrend":[{"scraped_at":(now-datetime.timedelta(hours=23-i)).isoformat(),"demand_mw":nm+2000+rnd.randint(-3000,3000),"generation_mw":nm+4500+rnd.randint(-2000,2000)} for i in range(24)],
        "_dbStats":get_db_stats(),
        "_sldcData":[{"sldc":s["name"],"state":s["state"],"region":s["region"],
            "demand_mw":int({"Delhi":5800,"Maharashtra":21000,"Gujarat":16000,"Tamil Nadu":14000,"Uttar Pradesh":22000,"Rajasthan":12000,"West Bengal":9000,"Karnataka":11000,"Andhra Pradesh":10000,"Madhya Pradesh":10500,"Punjab":7000,"Haryana":6000,"Odisha":4500,"Kerala":4800,"Bihar":6000}.get(s["state"],rnd.randint(500,5000))*rnd.uniform(.92,1.08)),
            "generation_mw":int({"Delhi":5800,"Maharashtra":21000,"Gujarat":16000,"Tamil Nadu":14000,"Uttar Pradesh":22000,"Rajasthan":12000,"West Bengal":9000,"Karnataka":11000,"Andhra Pradesh":10000,"Madhya Pradesh":10500,"Punjab":7000,"Haryana":6000,"Odisha":4500,"Kerala":4800,"Bihar":6000}.get(s["state"],rnd.randint(500,5000))*rnd.uniform(.88,1.05)),
            "freq_hz":round(50+rnd.uniform(-.12,.12),3),
            "peak_demand_mw":int({"Delhi":5800,"Maharashtra":21000,"Gujarat":16000,"Tamil Nadu":14000,"Uttar Pradesh":22000}.get(s["state"],rnd.randint(500,5000))*1.2)
            } for s in SLDCS],
        "_futuresData":[
            {"exchange":"NSE","symbol":"ELECMBL","expiry_date":"25-Jun-2026","last_price":4368,"open_price":4430,"high_price":4430,"low_price":4301,"close_price":4364,"prev_close":4350,"change_rs":14,"change_pct":0.32,"volume_lots":6863,"volume_mwh":343150000,"open_interest":4200,"turnover_cr":149.81,"vwap":4368,"unit":"₹/MWh"},
            {"exchange":"NSE","symbol":"ELECMBL","expiry_date":"25-Jul-2026","last_price":4410,"open_price":4380,"high_price":4450,"low_price":4360,"close_price":4410,"prev_close":4390,"change_rs":20,"change_pct":0.46,"volume_lots":2100,"volume_mwh":105000000,"open_interest":1800,"turnover_cr":46.3,"vwap":4405,"unit":"₹/MWh"},
            {"exchange":"MCX","symbol":"MCX-ELEC","expiry_date":"30-Jun-2026","last_price":4355,"open_price":4400,"high_price":4420,"low_price":4290,"close_price":4355,"prev_close":4340,"change_rs":15,"change_pct":0.35,"volume_lots":3200,"volume_mwh":160000000,"open_interest":2900,"turnover_cr":69.6,"vwap":4358,"unit":"₹/MWh"},
        ],
    }

# ── Cache & server ────────────────────────────────────────────────────────────
_cache={"data":None,"ts":0}; _lock=threading.Lock()

def cached_data(force=False):
    with _lock:
        if force or (time.time()-_cache["ts"]>CACHE_TTL):
            d=scrape_all()
            d["_freqTrend"]    =get_freq_trend(24)
            d["_mcpTrend"]     =get_mcp_trend(30)
            d["_stateData"]    =get_state_demand_latest()
            d["_regionalData"] =get_regional_latest()
            d["_rldcData"]     =get_rldc_latest()
            d["_nationalTrend"]=get_national_trend(720)  # 30 days for ratio chart
            d["_dbStats"]      =get_db_stats()
            d["_sldcData"]     =get_sldc_latest()
            d["_futuresData"]  =get_futures_latest()
            d["_futuresTrend"] =get_futures_trend(30)
            d["_peakNational"] =get_peak_demand_national(30)
            d["_peakStates"]   =get_peak_demand_states(30)
            d["_peakYearly"]   =get_peak_yearly_for_dashboard()
            d["_curtailment"]  =get_curtailment_data(60)
            d["_reCapacity"]   =get_re_capacity()
            d["_storageData"]  =get_storage_data()
            d["_goaData"]      =get_goa_data()
            d["_goaHistory"]   =get_goa_history(30)
            d["_monthlyVolume"]=_monthly_vol()
            _cache["data"]=d; _cache["ts"]=time.time()
        return _cache["data"]

def _monthly_vol():
    con=db()
    rows=con.execute("SELECT strftime('%Y-%m',date) mo, AVG(mcv_mwh)/1000 dam_mu FROM iex_dam_summary WHERE date IS NOT NULL GROUP BY mo ORDER BY mo DESC LIMIT 6").fetchall()
    con.close()
    if not rows: return {"labels":["Dec","Jan","Feb","Mar","Apr","May"],"dam":[9140,9810,8320,10450,10920,11380],"rtm":[3210,3480,2980,3890,4120,4310]}
    labels=[r[0] for r in reversed(rows)]; dam=[round(r[1] or 0,1) for r in reversed(rows)]
    return {"labels":labels,"dam":dam,"rtm":[round(x*.34,1) for x in dam]}

class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        p=urlparse(self.path); qs=parse_qs(p.query); path=p.path
        routes={
            "/data":           lambda: cached_data(),
            "/demo":           lambda: build_demo(),
            "/refresh":        lambda: (cached_data(force=True),{"status":"ok","ts":ts()})[1],
            "/dbstats":        lambda: get_db_stats(),
            "/log":            lambda: get_scrape_log(),
            "/trend/mcp":      lambda: get_mcp_trend(int(qs.get("days",["30"])[0])),
            "/trend/freq":     lambda: get_freq_trend(int(qs.get("hours",["24"])[0])),
            "/trend/regional": lambda: get_regional_trend(int(qs.get("days",["30"])[0])),
            "/trend/national": lambda: get_national_trend(int(qs.get("hours",["48"])[0])),
            "/states/latest":  lambda: get_state_demand_latest(),
            "/regional/latest":lambda: get_regional_latest(),
            "/rldc/latest":    lambda: get_rldc_latest(),
            "/sldc/latest":    lambda: get_sldc_latest(),
            "/futures/latest": lambda: get_futures_latest(),
            "/futures/trend":  lambda: get_futures_trend(int(qs.get("days",["30"])[0])),
            "/peak/national":  lambda: get_peak_demand_national(int(qs.get("days",["30"])[0])),
            "/peak/states":    lambda: get_peak_demand_states(int(qs.get("days",["30"])[0])),
            "/peak/yearly":    lambda: get_peak_demand_yearly(int(qs.get("year",[str(datetime.date.today().year)])[0])),
            "/curtailment":    lambda: get_curtailment_data(int(qs.get("days",["60"])[0])),
            "/re/capacity":    lambda: get_re_capacity(),
            "/storage":        lambda: get_storage_data(),
            "/goa":            lambda: get_goa_data(),
            "/goa/history":    lambda: get_goa_history(int(qs.get("days",["30"])[0])),
        }
        if path in routes: self._json(routes[path]())
        elif path=="/history": self._json({"table":qs.get("table",["state_demand"])[0],"rows":get_history(qs.get("table",["state_demand"])[0],int(qs.get("days",["30"])[0]))})
        elif path=="/state":   self._json(get_state_trend(qs.get("name",["Maharashtra"])[0],int(qs.get("days",["30"])[0])))
        elif path=="/export/excel": self._excel()
        elif path in ("/","dashboard.html") or path.endswith(".html"): self._file(DASH_PATH,"text/html")
        else: self._err(404,"Not found")

    def _json(self,obj):
        body=json.dumps(obj,default=str).encode()
        try:
            self.send_response(200); self.send_header("Content-Type","application/json")
            self.send_header("Access-Control-Allow-Origin","*"); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)
        except (ConnectionAbortedError, BrokenPipeError, ConnectionResetError): pass

    def _excel(self):
        export_excel()
        with open(EXCEL_PATH,"rb") as f: body=f.read()
        fname=f"india_power_{datetime.date.today().isoformat()}.xlsx"
        try:
            self.send_response(200); self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",f'attachment; filename="{fname}"')
            self.send_header("Content-Length",str(len(body))); self.send_header("Access-Control-Allow-Origin","*"); self.end_headers(); self.wfile.write(body)
        except (ConnectionAbortedError, BrokenPipeError, ConnectionResetError): pass

    def _file(self,path,mime):
        if not path.exists(): self._err(404,"File not found"); return
        body=path.read_bytes()
        try:
            self.send_response(200)
            self.send_header("Content-Type",mime); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)
        except (ConnectionAbortedError, BrokenPipeError, ConnectionResetError): pass

    def _err(self,code,msg):
        body=json.dumps({"error":msg}).encode(); self.send_response(code)
        self.send_header("Content-Type","application/json"); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)

    def log_message(self,fmt,*a): print(f"[{ts()[:19]}] {fmt%a}",file=sys.stderr)

def export_static_json(output_path=None):
    """
    Export a complete data snapshot as data.json for GitHub Pages deployment.
    This reads from the local SQLite DB (populated by the live scraper)
    and writes a single JSON file that the static dashboard can fetch.
    """
    if output_path is None:
        output_path = BASE / "data.json"

    print(f"[export-json] Building data snapshot...", file=sys.stderr)
    d = {}

    # Pull latest data from DB — same as the live /data endpoint
    # but WITHOUT triggering a new scrape (reads existing DB only)
    try:
        from scraper_v4 import (get_state_demand_latest, get_regional_latest,
                                 get_rldc_latest, get_national_trend,
                                 get_db_stats, get_sldc_latest,
                                 get_futures_latest, get_futures_trend,
                                 get_peak_demand_national, get_peak_demand_states,
                                 get_curtailment_data, get_re_capacity,
                                 get_storage_data, get_goa_data, get_goa_history,
                                 get_freq_trend, get_mcp_trend, _monthly_vol)
    except ImportError:
        pass  # running as __main__, functions are already in scope

    # Latest scrape result from DB (most recent rows, no new HTTP calls)
    now_ts = datetime.datetime.now().isoformat()

    # IEX DAM — latest day
    try:
        con = db()
        iex_row = con.execute(
            "SELECT * FROM iex_dam_summary ORDER BY date DESC, scraped_at DESC LIMIT 1"
        ).fetchone()
        iex_cols = [c[0] for c in con.execute("PRAGMA table_info(iex_dam_summary)").fetchall()]
        iex_latest = dict(zip(iex_cols, iex_row)) if iex_row else {}

        # DAM blocks for today
        blocks_rows = con.execute(
            "SELECT block_no,time_slot,mcp,purchase_mw,sell_mw FROM iex_dam_blocks "
            "WHERE date=(SELECT MAX(date) FROM iex_dam_blocks) ORDER BY CAST(block_no AS INTEGER)"
        ).fetchall()
        blocks = [{"block": r[0], "time": r[1], "mcp": r[2],
                   "purchase_mw": r[3], "sell_mw": r[4]} for r in blocks_rows]

        # POSOCO national latest
        nat_row = con.execute(
            "SELECT * FROM national_demand ORDER BY scraped_at DESC LIMIT 1"
        ).fetchone()
        nat_cols = [c[0] for c in con.execute("PRAGMA table_info(national_demand)").fetchall()]
        nat = dict(zip(nat_cols, nat_row)) if nat_row else {}

        con.close()
    except Exception as e:
        print(f"  DB read error: {e}", file=sys.stderr)
        iex_latest = {}; blocks = []; nat = {}

    d["scraped_at"]     = now_ts
    d["is_snapshot"]    = True          # tells dashboard this is a static snapshot
    d["snapshot_date"]  = datetime.date.today().isoformat()
    d["snapshot_time"]  = datetime.datetime.now().strftime("%H:%M IST")

    d["iex_dam"]        = {**iex_latest, "blocks": blocks,
                           "status": "ok" if iex_latest else "no_data"}
    d["posoco"]         = {"national": nat, "status": "ok" if nat else "no_data"}
    d["_national"]      = {"demand_mw": nat.get("demand_mw"),
                           "generation_mw": nat.get("generation_mw"),
                           "deficit_mw": nat.get("deficit_mw"),
                           "surplus_mw": nat.get("surplus_mw"),
                           "freq": nat.get("grid_freq_hz")}

    # All the trend / detail tables
    d["merit"]          = {"status": "ok",
                           "generation_table": _get_merit_latest()}
    d["_freqTrend"]     = get_freq_trend(24)
    d["_mcpTrend"]      = get_mcp_trend(30)
    d["_stateData"]     = get_state_demand_latest()
    d["_regionalData"]  = get_regional_latest()
    d["_rldcData"]      = get_rldc_latest()
    d["_nationalTrend"] = get_national_trend(48)
    d["_dbStats"]       = get_db_stats()
    d["_sldcData"]      = get_sldc_latest()
    d["_futuresData"]   = get_futures_latest()
    d["_futuresTrend"]  = get_futures_trend(30)
    d["_peakNational"]  = get_peak_demand_national(30)
    d["_peakStates"]    = get_peak_demand_states(30)
    d["_peakYearly"]    = get_peak_yearly_for_dashboard()
    d["_curtailment"]   = get_curtailment_data(60)
    d["_reCapacity"]    = get_re_capacity()
    d["_storageData"]   = get_storage_data()
    d["_goaData"]       = get_goa_data()
    d["_goaHistory"]    = get_goa_history(30)
    d["_monthlyVolume"] = _monthly_vol()

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(d, f, default=str, ensure_ascii=False)

    size_kb = pathlib.Path(output_path).stat().st_size // 1024
    print(f"[export-json] ✓ Written to {output_path} ({size_kb} KB)", file=sys.stderr)
    return str(output_path)


def _get_merit_latest():
    """Latest MERIT generation table from DB."""
    try:
        con = db()
        rows = con.execute(
            "SELECT source, generation_mw FROM merit_generation "
            "WHERE scraped_at=(SELECT MAX(scraped_at) FROM merit_generation) "
            "ORDER BY generation_mw DESC"
        ).fetchall()
        con.close()
        return [{"source": r[0], "generation_mw": r[1]} for r in rows]
    except Exception:
        return []


def push_to_github(repo_path, commit_message=None):
    """
    Commit data.json to the local GitHub repo clone and push.
    repo_path: path to your local clone of the GitHub Pages repo.
    """
    import subprocess, shutil

    repo = pathlib.Path(repo_path)
    if not repo.exists():
        print(f"[push] ERROR: repo path {repo_path} does not exist", file=sys.stderr)
        return False

    # Copy data.json to the repo
    src = BASE / "data.json"
    dst = repo / "data.json"
    shutil.copy2(src, dst)
    print(f"[push] Copied {src} → {dst}", file=sys.stderr)

    msg = commit_message or f"Auto-update {datetime.datetime.now().strftime('%Y-%m-%d %H:%M IST')}"

    try:
        def run(cmd):
            result = subprocess.run(cmd, cwd=repo, capture_output=True, text=True)
            if result.returncode != 0:
                print(f"[push] git error: {result.stderr}", file=sys.stderr)
                return False
            return True

        run(["git", "add", "data.json"])
        # Check if there's anything to commit
        status = subprocess.run(["git", "status", "--porcelain"],
                                 cwd=repo, capture_output=True, text=True)
        if not status.stdout.strip():
            print("[push] No changes to commit (data.json unchanged)", file=sys.stderr)
            return True

        run(["git", "commit", "-m", msg])
        run(["git", "push", "origin", "main"])
        print(f"[push] ✓ Pushed: {msg}", file=sys.stderr)
        return True
    except FileNotFoundError:
        print("[push] ERROR: git not found. Install git and ensure it's in PATH.", file=sys.stderr)
        return False
    except Exception as e:
        print(f"[push] ERROR: {e}", file=sys.stderr)
        return False


def start_scheduler(interval):
    def run():
        while True:
            time.sleep(interval*60)
            try: cached_data(force=True)
            except Exception as e: print(f"Scheduler error: {e}",file=sys.stderr)
    threading.Thread(target=run,daemon=True).start()
    print(f"Auto-refresh every {interval} min",file=sys.stderr)

if __name__=="__main__":
    parser=argparse.ArgumentParser(
        description="India Power Market Scraper v5.3",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scraper_v4.py --serve 8080 --interval 15   # Live server (your PC)
  python scraper_v4.py --scrape                      # One scrape then exit
  python scraper_v4.py --export                      # Export Excel
  python scraper_v4.py --export-json                 # Export data.json for GitHub Pages
  python scraper_v4.py --export-json --push-github /path/to/repo  # Export + auto git push
        """
    )
    parser.add_argument("--serve",        type=int,  metavar="PORT",
                        help="Start live server on PORT (also starts 15-min scraper)")
    parser.add_argument("--scrape",       action="store_true",
                        help="Run one full scrape cycle and exit")
    parser.add_argument("--export",       action="store_true",
                        help="Export all data to Excel (.xlsx)")
    parser.add_argument("--interval",     type=int,  default=15,
                        help="Scrape interval in minutes when using --serve (default: 15)")
    parser.add_argument("--export-json",  action="store_true",
                        help="Export data snapshot as data.json for GitHub Pages")
    parser.add_argument("--push-github",  metavar="REPO_PATH",
                        help="After --export-json, git-commit and push to this local repo clone")
    parser.add_argument("--json-out",     metavar="PATH", default=None,
                        help="Output path for --export-json (default: same folder as scraper)")
    args=parser.parse_args()
    init_db()

    if args.scrape:
        d=scrape_all(); print(json.dumps({"status":"done","scraped_at":d["scraped_at"]},indent=2))

    elif args.export:
        export_excel()

    elif args.export_json:
        out = export_static_json(args.json_out)
        if args.push_github:
            push_to_github(args.push_github)

    elif args.serve:
        start_scheduler(args.interval)
        print(f"\n⚡ India Power Market Server v5.3"
              f"\n   Dashboard   : http://localhost:{args.serve}/"
              f"\n   Live data   : http://localhost:{args.serve}/data"
              f"\n   RLDC latest : http://localhost:{args.serve}/rldc/latest"
              f"\n   States      : http://localhost:{args.serve}/states/latest"
              f"\n   Excel       : http://localhost:{args.serve}/export/excel"
              f"\n   SLDC latest : http://localhost:{args.serve}/sldc/latest"
              f"\n   Futures     : http://localhost:{args.serve}/futures/latest"
              f"\n   Log         : http://localhost:{args.serve}/log\n", file=sys.stderr)
        HTTPServer(("",args.serve),Handler).serve_forever()

    else:
        parser.print_help()
